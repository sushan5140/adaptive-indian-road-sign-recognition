"""Package the metadata-only next phase of unseen-class acquisition planning."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import sys
import tempfile
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from data.acquisition_planning import (  # noqa: E402
    AcquisitionClassState,
    AcquisitionPlanningError,
    build_acquisition_plan,
)

OUTPUT_DIR = PROJECT_ROOT / "outputs/open_set_acquisition"
PDF_PATH = OUTPUT_DIR / "acquisition_next_phase_report.pdf"
XLSX_PATH = OUTPUT_DIR / "acquisition_next_phase_summary.xlsx"
CSV_PATH = OUTPUT_DIR / "acquisition_next_phase_summary.csv"
IRSDB_DRAFT = OUTPUT_DIR / "irsdb_access_request.md"
DATASET_B_DRAFT = OUTPUT_DIR / "dataset_b_licence_request.md"

MAPILLARY_STATUS = "not_queried_authentication_unavailable"
MAPILLARY_CANDIDATE_COLUMNS = (
    "mapillary_image_id",
    "sequence_id",
    "capture_time",
    "approximate_location",
    "contributor_id_or_name",
    "traffic_sign_class_or_detection_label",
    "source_url_or_reference",
    "licence_terms_snapshot_or_citation",
    "proposed_project_class",
    "class_mapping_confidence",
    "dependency_group",
)

SOURCE_URLS = {
    "mapillary_features": "https://help.mapillary.com/hc/en-us/articles/115002332165-Map-features",
    "mapillary_download": "https://help.mapillary.com/hc/en-us/articles/4407521157138-Downloading-map-data-via-the-Mapillary-web-app",
    "mapillary_licence": "https://help.mapillary.com/hc/en-us/articles/115001770409-CC-BY-SA-license-for-open-data",
    "mapillary_terms": "https://www.mapillary.com/terms",
    "irsdb": "https://www.nitrkl.ac.in/docs/CS/Database/dataset_writeup.pdf",
    "irsdb_publication": "https://doi.org/10.1007/978-981-15-4032-5_63",
    "dataset_b": "https://doi.org/10.5281/zenodo.17300841",
    "gtsrb": "https://benchmark.ini.rub.de/gtsrb_dataset.html",
    "gtsrb_paper": "https://www.ini.rub.de/upload/file/1470692859_c57fac98ca9d02ac701c/stallkampetal_gtsrb_nn_si2012.pdf",
}


def main() -> int:
    """Generate and validate the next-phase acquisition evidence package."""
    try:
        _validate_inputs_and_destinations()
        current_status = _current_status_rows()
        mapillary_candidates: list[dict[str, str]] = []
        irsdb_rows = _irsdb_rows()
        dataset_b_rows = _dataset_b_rows()
        gtsrb_rows = _gtsrb_rows()
        plan_rows = _acquisition_plan_rows()
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(
            prefix="acquisition-next-phase-", dir=OUTPUT_DIR
        ) as temporary_directory:
            stage = Path(temporary_directory)
            staged_pdf = stage / PDF_PATH.name
            staged_xlsx = stage / XLSX_PATH.name
            staged_csv = stage / CSV_PATH.name
            _write_csv(staged_csv, plan_rows)
            _write_workbook(
                staged_xlsx,
                current_status,
                mapillary_candidates,
                irsdb_rows,
                dataset_b_rows,
                gtsrb_rows,
                plan_rows,
            )
            _write_pdf(
                staged_pdf,
                current_status,
                irsdb_rows,
                dataset_b_rows,
                gtsrb_rows,
                plan_rows,
            )
            page_count = _validate_outputs(
                staged_pdf, staged_xlsx, staged_csv, plan_rows
            )
            os.replace(staged_pdf, PDF_PATH)
            os.replace(staged_xlsx, XLSX_PATH)
            os.replace(staged_csv, CSV_PATH)
        result = {
            "mapillary_authentication_worked": False,
            "mapillary_metadata_records_collected": 0,
            "mapillary_candidate_counts": {
                class_name: 0
                for class_name in (
                    "stop",
                    "maximum_speed_limit_50_km_h",
                    "no_left_turn",
                    "no_parking",
                )
            },
            "candidate_count_interpretation": (
                "zero records collected because authentication was unavailable; "
                "this is not evidence that Mapillary has zero matching signs"
            ),
            "pdf_pages": page_count,
            "xlsx_sheets": [
                "Current Status",
                "Mapillary Candidates",
                "IRSDB",
                "Dataset B Licence",
                "GTSRB Fallback",
                "Acquisition Plan",
            ],
            "image_pixels_downloaded": 0,
            "requests_sent": 0,
            "training_prototype_calibration_or_evaluation_performed": False,
            "artifacts": {
                path.name: {"path": str(path), "sha256": _sha256(path)}
                for path in (
                    PDF_PATH,
                    XLSX_PATH,
                    CSV_PATH,
                    IRSDB_DRAFT,
                    DATASET_B_DRAFT,
                )
            },
        }
        print(json.dumps(result, indent=2))
        return 0
    except (
        AcquisitionPlanningError,
        KeyError,
        OSError,
        TypeError,
        ValueError,
    ) as error:
        print(f"FATAL: {error}", file=sys.stderr)
        return 2


def _validate_inputs_and_destinations() -> None:
    for draft in (IRSDB_DRAFT, DATASET_B_DRAFT):
        if not draft.is_file():
            raise AcquisitionPlanningError(f"Required draft is missing: {draft}")
    for destination in (PDF_PATH, XLSX_PATH, CSV_PATH):
        if destination.exists():
            raise AcquisitionPlanningError(
                f"Unique output name already exists; choose a new name: {destination}"
            )
    credential_names = (
        "MAPILLARY_ACCESS_TOKEN",
        "MAPILLARY_TOKEN",
        "MAPILLARY_CLIENT_TOKEN",
    )
    if any(os.getenv(name) for name in credential_names):
        raise AcquisitionPlanningError(
            "A Mapillary credential became available; run the authenticated "
            "metadata query before generating this no-auth report"
        )


def _current_status_rows() -> list[dict[str, Any]]:
    values = (
        ("stop", 0, 0, 0, "missing"),
        ("no_left_turn", 1, 1, 1, "licence_blocked_and_quantity_weak"),
        ("maximum_speed_limit_50_km_h", 0, 0, 0, "missing"),
        ("no_parking", 4, 3, 4, "licence_blocked_and_quantity_weak"),
        ("bus_stop", 34, 29, 34, "quantity_met_but_licence_blocked"),
    )
    return [
        {
            "class_name": label,
            "visually_accepted_photos": photos,
            "accepted_independent_groups": groups,
            "accepted_but_licence_blocked": blocked,
            "licence_ready_photos": 0,
            "current_status": status,
        }
        for label, photos, groups, blocked, status in values
    ]


def _irsdb_rows() -> list[dict[str, str]]:
    return [
        {
            "source_name": "IRSDBv1.0 Fully Annotated Indian Traffic Signs Database",
            "official_url": SOURCE_URLS["irsdb"],
            "institution": "NIT Rourkela Centre for Computer Vision & Pattern Recognition",
            "documented_images": "1692",
            "documented_classes": "49",
            "physical_sign_grouping": "training tracks represent one physical traffic sign",
            "access_terms": "request-only research use for individuals and non-commercial academic laboratories; citation required",
            "target_class_availability": "unknown until the 49-class manifest is supplied",
            "request_status": "draft prepared; not sent",
            "draft_path": IRSDB_DRAFT.relative_to(PROJECT_ROOT).as_posix(),
        }
    ]


def _dataset_b_rows() -> list[dict[str, str]]:
    return [
        {
            "source_name": "Indian Traffic VQA Dataset B",
            "doi": "10.5281/zenodo.17300841",
            "official_url": SOURCE_URLS["dataset_b"],
            "documented_images": "1085 real-world Indian photographs",
            "rights_record": "Zenodo Rights/License field is blank in the inspected record",
            "affected_candidates": "39 visually accepted: bus_stop 34; no_left_turn 1; no_parking 4",
            "licence_status": "pending_curator_confirmation",
            "experiment_ready_candidates": "0",
            "request_status": "draft prepared; not sent",
            "draft_path": DATASET_B_DRAFT.relative_to(PROJECT_ROOT).as_posix(),
        }
    ]


def _gtsrb_rows() -> list[dict[str, str]]:
    common = {
        "official_url": SOURCE_URLS["gtsrb"],
        "geographic_domain": "Germany; cross-domain only",
        "reuse_terms": "official site states data is free to use and requests citation",
        "track_structure": "30 dependent frames per unique physical-sign track",
        "track_count": "not measured; official archive was not downloaded in this phase",
        "download_status": "not downloaded",
        "scientific_limitation": "German sign/domain; cropped sign tracks cannot support Indian-domain performance claims",
    }
    return [
        {
            "project_class": "stop",
            "gtsrb_class_id": "14",
            "class_mapping": "exact semantic class",
            "fallback_suitability": "conditional; verify at least 15 training tracks after deliberate acquisition",
            **common,
        },
        {
            "project_class": "maximum_speed_limit_50_km_h",
            "gtsrb_class_id": "2",
            "class_mapping": "exact numeral and semantic class",
            "fallback_suitability": "conditional; verify at least 15 training tracks after deliberate acquisition",
            **common,
        },
    ]


def _acquisition_plan_rows() -> list[dict[str, Any]]:
    states = (
        AcquisitionClassState("stop", 0, 0),
        AcquisitionClassState("no_left_turn", 1, 1),
        AcquisitionClassState("maximum_speed_limit_50_km_h", 0, 0),
        AcquisitionClassState("no_parking", 4, 3),
        AcquisitionClassState("bus_stop", 34, 29),
    )
    recommendations = {
        "stop": "Mapillary India after authentication plus IRSDB; GTSRB class 14 only as explicit cross-domain fallback",
        "no_left_turn": "Mapillary India after authentication plus IRSDB",
        "maximum_speed_limit_50_km_h": "Mapillary India after authentication plus IRSDB; GTSRB class 2 only as explicit cross-domain fallback",
        "no_parking": "Mapillary India after authentication plus IRSDB",
        "bus_stop": "No additional acquisition now; obtain Dataset B curator licence confirmation",
    }
    rows = build_acquisition_plan(
        states,
        mapillary_counts={state.class_name: 0 for state in states},
        mapillary_status=MAPILLARY_STATUS,
        source_recommendations=recommendations,
    )
    for row in rows:
        label = row["class_name"]
        row["irsdb_availability_status"] = (
            "not_required_for_current_quantity"
            if label == "bus_stop"
            else "unknown_pending_access_request_and_class_manifest"
        )
        row["gtsrb_fallback_availability"] = (
            "conditional_exact_class_archive_not_acquired"
            if label in {"stop", "maximum_speed_limit_50_km_h"}
            else "not_applicable"
        )
        if label == "bus_stop":
            row["mapillary_query_status"] = (
                "not_searched_additional_quantity_not_needed"
            )
    return rows


def _write_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    columns = tuple(rows[0])
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="raise")
        writer.writeheader()
        writer.writerows(rows)


def _write_workbook(
    path: Path,
    current_status: list[dict[str, Any]],
    mapillary_candidates: list[dict[str, str]],
    irsdb_rows: list[dict[str, str]],
    dataset_b_rows: list[dict[str, str]],
    gtsrb_rows: list[dict[str, str]],
    plan_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets: list[tuple[str, Sequence[str], Sequence[Mapping[str, Any]]]] = [
        ("Current Status", tuple(current_status[0]), current_status),
        ("Mapillary Candidates", MAPILLARY_CANDIDATE_COLUMNS, mapillary_candidates),
        ("IRSDB", tuple(irsdb_rows[0]), irsdb_rows),
        ("Dataset B Licence", tuple(dataset_b_rows[0]), dataset_b_rows),
        ("GTSRB Fallback", tuple(gtsrb_rows[0]), gtsrb_rows),
        ("Acquisition Plan", tuple(plan_rows[0]), plan_rows),
    ]
    for name, columns, rows in sheets:
        sheet = workbook.create_sheet(name)
        sheet.append(list(columns))
        for row in rows:
            sheet.append([row[column] for column in columns])
        _style_sheet(sheet, columns, rows)
    workbook.save(path)


def _style_sheet(
    sheet: Any, columns: Sequence[str], rows: Sequence[Mapping[str, Any]]
) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, 1):
        values = [str(column), *(str(row[column]) for row in rows)]
        width = min(max(len(value) for value in values) + 2, 65)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 12)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_pdf(
    path: Path,
    current_status: list[dict[str, Any]],
    irsdb_rows: list[dict[str, str]],
    dataset_b_rows: list[dict[str, str]],
    gtsrb_rows: list[dict[str, str]],
    plan_rows: list[dict[str, Any]],
) -> None:
    styles = getSampleStyleSheet()
    small = ParagraphStyle(
        "SmallEvidence",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        spaceAfter=2 * mm,
    )
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=11 * mm,
        rightMargin=11 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title="Open-set acquisition next-phase report",
        author="Adaptive Indian Road Sign Recognition project",
    )
    story: list[Any] = [
        Paragraph("Open-Set Acquisition: Next-Phase Report", styles["Title"]),
        Paragraph(
            "Evidence date: 2026-09-01. Metadata and permission planning only. No image pixels were downloaded; no request was sent; no model operation was performed.",
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
        _rows_table(
            current_status,
            (
                "class_name",
                "visually_accepted_photos",
                "accepted_independent_groups",
                "accepted_but_licence_blocked",
                "licence_ready_photos",
                "current_status",
            ),
            [45 * mm, 35 * mm, 40 * mm, 42 * mm, 34 * mm, 80 * mm],
            8,
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "Locked threshold: at least 30 approved photographs and 15 independent groups per class. Visual acceptance is not licence approval.",
            styles["BodyText"],
        ),
        PageBreak(),
        Paragraph("Mapillary metadata-only discovery", styles["Heading1"]),
        _key_value_table(
            [
                (
                    "Authentication",
                    "Unavailable: no Mapillary token or credential file was present.",
                ),
                ("Authenticated query", "Not run."),
                (
                    "Metadata candidates collected",
                    "0 for stop, speed-50, no-left-turn, and no-parking; this is an unqueried count, not evidence of absence.",
                ),
                ("Pixels downloaded", "0"),
                (
                    "Deduplication",
                    "Deferred until metadata exist; image ID, sequence ID, capture time/location dependency, and contributor evidence must be retained.",
                ),
                (
                    "Exact blocker",
                    "Create/register a Mapillary application, provide a valid access token outside source control, verify exact India taxonomy mappings, and snapshot current terms before querying.",
                ),
            ]
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "Official guidance says Mapillary exposes 1,500 traffic-sign map-feature classes and recommends API access for large areas. Traffic-sign features are derived from multiple detections, so multiple associated images must not be treated as independent. The download guide requires Terms compliance. The current Terms page was not readable without login. The official licence help page states imagery is CC BY-SA with per-image attribution.",
            small,
        ),
        Paragraph(SOURCE_URLS["mapillary_features"], small),
        Paragraph(SOURCE_URLS["mapillary_download"], small),
        Paragraph(SOURCE_URLS["mapillary_licence"], small),
        Paragraph(SOURCE_URLS["mapillary_terms"], small),
        PageBreak(),
        Paragraph("IRSDBv1.0 request preparation", styles["Heading1"]),
        _rows_table(
            irsdb_rows,
            (
                "source_name",
                "documented_images",
                "documented_classes",
                "physical_sign_grouping",
                "access_terms",
                "target_class_availability",
                "request_status",
            ),
            [50 * mm, 25 * mm, 25 * mm, 52 * mm, 60 * mm, 55 * mm, 30 * mm],
            7,
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "The draft requests access, the complete class manifest, per-class physical-track counts, exact target-class confirmation, and explicit permissions for training, crops, private storage, report examples, aggregate metrics, and redistribution limits.",
            styles["BodyText"],
        ),
        Paragraph(SOURCE_URLS["irsdb"], small),
        PageBreak(),
        Paragraph("Dataset B licence blocker", styles["Heading1"]),
        _rows_table(
            dataset_b_rows,
            (
                "source_name",
                "doi",
                "documented_images",
                "rights_record",
                "affected_candidates",
                "licence_status",
                "experiment_ready_candidates",
                "request_status",
            ),
            [42 * mm, 32 * mm, 45 * mm, 54 * mm, 55 * mm, 42 * mm, 28 * mm, 32 * mm],
            7,
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "Zenodo hosts the record but does not own uploaded content; permission depends on the record's Rights/License field. The inspected record shows the License heading without a value. The concise draft asks the listed curators to bind explicit terms to the exact traffic512final.zip files.",
            styles["BodyText"],
        ),
        Paragraph(SOURCE_URLS["dataset_b"], small),
        PageBreak(),
        Paragraph("GTSRB cross-domain fallback assessment", styles["Heading1"]),
        _rows_table(
            gtsrb_rows,
            (
                "project_class",
                "gtsrb_class_id",
                "class_mapping",
                "track_count",
                "fallback_suitability",
                "reuse_terms",
                "scientific_limitation",
            ),
            [42 * mm, 25 * mm, 42 * mm, 50 * mm, 60 * mm, 55 * mm, 60 * mm],
            7,
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "Official GTSRB documentation states that each training track contains 30 frames of one unique physical sign and that the data is free to use with citation requested. Class IDs 14 (STOP) and 2 (speed limit 50 km/h) map exactly. Per-class track counts were not measured because no archive was downloaded. Suitability remains conditional until at least 15 distinct training tracks are verified. Any result must be labelled German cross-domain evidence, never Indian-domain performance.",
            styles["BodyText"],
        ),
        Paragraph(SOURCE_URLS["gtsrb"], small),
        Paragraph(SOURCE_URLS["gtsrb_paper"], small),
        PageBreak(),
        Paragraph("Per-class acquisition plan", styles["Heading1"]),
        _rows_table(
            plan_rows,
            (
                "class_name",
                "visually_accepted_photos",
                "visually_accepted_independent_groups",
                "remaining_photo_quantity_gap",
                "remaining_group_gap",
                "mapillary_metadata_candidate_count",
                "irsdb_availability_status",
                "gtsrb_fallback_availability",
                "acquisition_readiness",
            ),
            [
                43 * mm,
                27 * mm,
                32 * mm,
                29 * mm,
                26 * mm,
                31 * mm,
                47 * mm,
                48 * mm,
                45 * mm,
            ],
            6.5,
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "Exact next action: send neither draft yet. First obtain a Mapillary application token through the normal authenticated flow and review the logged-in Terms. Then run an India metadata-only query for the four weak classes, retain all sequence/dependency metadata, and manually verify taxonomy mappings before any pixel-download decision. In parallel, the user may review and send the two prepared permission requests. No metadata candidate is an approved photograph.",
            styles["BodyText"],
        ),
    ]
    document.build(story)


def _rows_table(
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
    widths: Sequence[float],
    font_size: float,
) -> Table:
    data = [[_header(column) for column in columns]]
    data.extend([[str(row[column]) for column in columns] for row in rows])
    return _table(data, widths, font_size)


def _key_value_table(rows: Sequence[tuple[str, str]]) -> Table:
    return _table(
        [["Field", "Result"], *[list(row) for row in rows]], [55 * mm, 220 * mm], 8
    )


def _table(
    data: Sequence[Sequence[Any]], widths: Sequence[float], font_size: float
) -> Table:
    body = ParagraphStyle(
        f"TableBody{font_size}",
        fontName="Helvetica",
        fontSize=font_size,
        leading=font_size + 2,
    )
    header = ParagraphStyle(
        f"TableHeader{font_size}",
        parent=body,
        fontName="Helvetica-Bold",
        textColor=colors.white,
    )
    formatted = [
        [
            Paragraph(_escape(str(value)), header if row_index == 0 else body)
            for value in row
        ]
        for row_index, row in enumerate(data)
    ]
    table = Table(formatted, colWidths=list(widths), repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#EDF3F8")],
                ),
            ]
        )
    )
    return table


def _validate_outputs(
    pdf_path: Path,
    xlsx_path: Path,
    csv_path: Path,
    plan_rows: Sequence[Mapping[str, Any]],
) -> int:
    pages = len(PdfReader(str(pdf_path)).pages)
    if pages != 6:
        raise AcquisitionPlanningError(f"Expected 6 PDF pages, found {pages}")
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    expected_sheets = [
        "Current Status",
        "Mapillary Candidates",
        "IRSDB",
        "Dataset B Licence",
        "GTSRB Fallback",
        "Acquisition Plan",
    ]
    if workbook.sheetnames != expected_sheets:
        workbook.close()
        raise AcquisitionPlanningError(
            f"Unexpected workbook sheets: {workbook.sheetnames}"
        )
    if workbook["Mapillary Candidates"].max_row != 1:
        workbook.close()
        raise AcquisitionPlanningError(
            "Mapillary sheet must contain no fabricated candidate records"
        )
    if workbook["Acquisition Plan"].max_row != 6:
        workbook.close()
        raise AcquisitionPlanningError("Acquisition Plan must contain five classes")
    workbook.close()
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        rows = [dict(row) for row in csv.DictReader(handle)]
    if len(rows) != 5 or {row["class_name"] for row in rows} != {
        str(row["class_name"]) for row in plan_rows
    }:
        raise AcquisitionPlanningError("Summary CSV class coverage mismatch")
    if any(row["metadata_candidates_counted_as_approved"] != "no" for row in rows):
        raise AcquisitionPlanningError("Metadata candidates were counted as approved")
    return pages


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _escape(value: str) -> str:
    return value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _header(value: str) -> str:
    return value.replace("_", " ").title()


if __name__ == "__main__":
    raise SystemExit(main())
