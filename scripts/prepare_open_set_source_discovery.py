"""Package external-source research and local unseen-class recovery evidence."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import sys
import tempfile
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PilImage
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from data.source_discovery import (  # noqa: E402
    SourceDiscoveryError,
    build_perceptual_groups,
    eligible_counts_by_class,
    frozen_split_overlap,
    make_split_evidence,
)

OUTPUT = PROJECT_ROOT / "outputs/open_set_acquisition"
PDF_PATH = OUTPUT / "source_discovery_report.pdf"
XLSX_PATH = OUTPUT / "source_discovery_summary.xlsx"
CSV_PATH = OUTPUT / "source_discovery_summary.csv"
VQA_IMAGE_ROOT = PROJECT_ROOT / "data/raw/indian_traffic_vqa/traffic512final"
DATASET_A_IMAGE_ROOT = (
    PROJECT_ROOT
    / "data/raw/indian_traffic_sign_dataset/Indian-Traffic Sign-Dataset/Images"
)

TARGET_CLASSES = (
    "stop",
    "no_left_turn",
    "maximum_speed_limit_50_km_h",
    "no_parking",
    "bus_stop",
)

DATASET_SOURCES: tuple[dict[str, str], ...] = (
    {
        "source_name": "Indian Traffic VQA Dataset (local Dataset B)",
        "official_url": "https://doi.org/10.5281/zenodo.17300841",
        "institution_maintainer": "Bapatla Engineering College contributors / Zenodo",
        "geographic_domain": "India",
        "relevant_classes": "no_left_turn; no_parking; annotations mention stop and speed 50 but visual audit disproves those rows",
        "documented_size": "1,085 images; 4,341 questions",
        "real_photographs": "yes",
        "approx_usable_for_targets": "2 no-left-turn + 5 no-parking images before new review",
        "license_research_terms": "Zenodo Rights field is blank. A matching maintainer Hugging Face card says non-commercial research use, modification, distribution, and attribution, but also carries an MIT metadata tag.",
        "repository_redistribution": "conditional-no until curator confirms the Hugging Face terms cover the exact Zenodo files",
        "legal_confidence": "medium-low",
        "scientific_confidence": "medium after per-image relabel review",
    },
    {
        "source_name": "IRSDBv1.0 Fully Annotated Indian Traffic Signs Database",
        "official_url": "https://www.nitrkl.ac.in/docs/CS/Database/dataset_writeup.pdf",
        "institution_maintainer": "NIT Rourkela, Centre for Computer Vision & Pattern Recognition",
        "geographic_domain": "India",
        "relevant_classes": "public write-up does not expose the 49-class manifest; request it before assuming coverage",
        "documented_size": "1,692 images; 49 classes; training tracks represent physical signs",
        "real_photographs": "yes",
        "approx_usable_for_targets": "unknown until access and class manifest are supplied",
        "license_research_terms": "Accessible on request for research use by individuals and non-commercial academic laboratories; citation required.",
        "repository_redistribution": "not granted by the public terms; request explicit permission",
        "legal_confidence": "high that research access is restricted; low for redistribution",
        "scientific_confidence": "high if target classes and independent tracks are confirmed",
    },
    {
        "source_name": "Mapillary platform imagery and traffic-sign map features",
        "official_url": "https://help.mapillary.com/hc/en-us/articles/4407521157138-Downloading-map-data-via-the-Mapillary-web-app",
        "institution_maintainer": "Mapillary / Meta",
        "geographic_domain": "Global; query India explicitly",
        "relevant_classes": "1,500 traffic-sign classes; exact India counts for all five targets require an authenticated metadata query",
        "documented_size": "1,500 traffic-sign classes; image count is query-dependent",
        "real_photographs": "yes",
        "approx_usable_for_targets": "unknown before India-region metadata query and manual validation",
        "license_research_terms": "Official help says images are shared under CC BY-SA with attribution; API/map-data use must also comply with Terms and Commercial Terms.",
        "repository_redistribution": "conditional yes for selected attributed images if current terms and share-alike obligations are followed",
        "legal_confidence": "medium-high; record per-image attribution and terms snapshot",
        "scientific_confidence": "high potential after exact-class and source-sequence grouping",
    },
    {
        "source_name": "Mapillary Traffic Sign Dataset (MTSD)",
        "official_url": "https://arxiv.org/abs/1909.04422",
        "institution_maintainer": "Mapillary Research",
        "geographic_domain": "Global",
        "relevant_classes": "313 fine-grained classes; exact target labels/counts require taxonomy inspection after accepting the licence",
        "documented_size": "about 100k images; 52,453 fully annotated split images; 47,547 partial; 257,543 boxes",
        "real_photographs": "yes",
        "approx_usable_for_targets": "not documented per target on the public paper page",
        "license_research_terms": "Research-use licence is non-transferable and bars derivatives and display/distribution of the dataset.",
        "repository_redistribution": "no",
        "legal_confidence": "high; restrictions are explicit",
        "scientific_confidence": "high for internal research, unsuitable for repository-packaged image redistribution",
    },
    {
        "source_name": "German Traffic Sign Recognition Benchmark (GTSRB)",
        "official_url": "https://benchmark.ini.rub.de/gtsrb_dataset.html",
        "institution_maintainer": "Institut fuer Neuroinformatik, Ruhr University Bochum",
        "geographic_domain": "Germany",
        "relevant_classes": "stop; maximum_speed_limit_50_km_h; roundabout mandatory is not roundabout-ahead",
        "documented_size": "more than 50,000 images; more than 40 classes; 30 frames per physical-sign track",
        "real_photographs": "yes",
        "approx_usable_for_targets": "per-class image and independent-track counts not stated on official overview",
        "license_research_terms": "Official page states the data is free to use and requests citation.",
        "repository_redistribution": "likely permissible under the official statement; preserve citation and verify archive readme before copying",
        "legal_confidence": "medium-high",
        "scientific_confidence": "high as cross-domain fallback; low for India-domain claims",
    },
    {
        "source_name": "LISA Traffic Signs Dataset",
        "official_url": "https://cvrr.ucsd.edu/lisa-traffic-signs-dataset",
        "institution_maintainer": "UC San Diego CVRR / LISA",
        "geographic_domain": "United States",
        "relevant_classes": "associated taxonomy includes stop, no-left-turn, and no-parking; verify exact archive counts",
        "documented_size": "47 sign types; 7,855 annotations on 6,610 frames",
        "real_photographs": "yes; annotated video frames",
        "approx_usable_for_targets": "not documented per target on official page",
        "license_research_terms": "Academic licence agreement; citation required.",
        "repository_redistribution": "not established by the public page; obtain/inspect agreement first",
        "legal_confidence": "medium-low until agreement is reviewed",
        "scientific_confidence": "medium; cross-domain and sequence dependence must be grouped",
    },
    {
        "source_name": "DATS_2022",
        "official_url": "https://doi.org/10.17632/nfc34n8svj.2",
        "institution_maintainer": "Bhakti Paranjape / Mendeley Data",
        "geographic_domain": "India",
        "relevant_classes": "only coarse/sparse traffic-object annotations are documented; target sign classes are not confirmed",
        "documented_size": "more than 10,000 images from photos and extracted video frames",
        "real_photographs": "yes",
        "approx_usable_for_targets": "unknown; would require a metadata-first annotation scan and human review",
        "license_research_terms": "CC BY 4.0",
        "repository_redistribution": "yes with attribution if target images are actually present",
        "legal_confidence": "high",
        "scientific_confidence": "low-medium because fine-grained sign labels and frame independence are absent",
    },
    {
        "source_name": "Mapillary Vistas Validation fine-grained relabeling (MVV)",
        "official_url": "https://github.com/nec-labs-ma/relabeling",
        "institution_maintainer": "NEC Laboratories America",
        "geographic_domain": "Global",
        "relevant_classes": "stop; speed-limit generic; no-parking; turn generic; roundabout nearly absent",
        "documented_size": "2,000 curated validation images; 11/12 coarse fine-sign categories",
        "real_photographs": "yes",
        "approx_usable_for_targets": "paper chart suggests about 220 stop and 500 no-parking annotations; exact filtered image counts require archive inspection",
        "license_research_terms": "Repository states research/academic use only and non-commercial use; underlying Mapillary rights still apply.",
        "repository_redistribution": "unclear for the underlying images; do not copy until confirmed",
        "legal_confidence": "medium-low",
        "scientific_confidence": "medium for stop/no-parking; labels too coarse for speed-50/no-left-turn",
    },
    {
        "source_name": "Belgium Traffic Sign Dataset (BelgiumTS/BTSC)",
        "official_url": "https://btsd.ethz.ch/shareddata/",
        "institution_maintainer": "ETH Zurich / KU Leuven, Radu Timofte",
        "geographic_domain": "Belgium",
        "relevant_classes": "62-sign reduced classification set; exact target mapping/counts require taxonomy inspection",
        "documented_size": "62 reduced sign types; classification train/test archives and multi-camera sequences",
        "real_photographs": "yes",
        "approx_usable_for_targets": "not documented per target on official page",
        "license_research_terms": "Official page requests publication citation but does not state an explicit data licence.",
        "repository_redistribution": "no until maintainer permission or explicit licence is obtained",
        "legal_confidence": "low",
        "scientific_confidence": "medium as cross-domain source with strong sequence grouping",
    },
    {
        "source_name": "Tsinghua-Tencent 100K (TT100K)",
        "official_url": "https://cg.cs.tsinghua.edu.cn/ctsdb/",
        "institution_maintainer": "Tsinghua University / Tencent",
        "geographic_domain": "China",
        "relevant_classes": "no-parking and speed-limit variants are documented by downstream taxonomies; exact five-class mapping requires official annotation inspection",
        "documented_size": "100k road images; 221 configured sign classes in common tooling",
        "real_photographs": "yes",
        "approx_usable_for_targets": "not documented per target on official page",
        "license_research_terms": "No explicit dataset licence was found on the official landing page.",
        "repository_redistribution": "no",
        "legal_confidence": "low",
        "scientific_confidence": "medium as cross-domain fallback if rights are clarified",
    },
)

LICENSE_URLS = {
    "Indian Traffic VQA Dataset (local Dataset B)": "https://huggingface.co/datasets/chandrabhuma/Indian_Traffic_VQA_Dataset",
    "IRSDBv1.0 Fully Annotated Indian Traffic Signs Database": "https://www.nitrkl.ac.in/docs/CS/Database/dataset_writeup.pdf",
    "Mapillary platform imagery and traffic-sign map features": "https://help.mapillary.com/hc/en-us/articles/115001770409-CC-BY-SA-license-for-open-data",
    "Mapillary Traffic Sign Dataset (MTSD)": "https://www.mapillary.com/dataset/assets/mapillary-object-dataset-research-use-license-2019.pdf",
    "German Traffic Sign Recognition Benchmark (GTSRB)": "https://benchmark.ini.rub.de/gtsrb_dataset.html",
    "LISA Traffic Signs Dataset": "https://cvrr.ucsd.edu/lisa-traffic-signs-dataset",
    "DATS_2022": "https://data.mendeley.com/datasets/nfc34n8svj/2",
    "Mapillary Vistas Validation fine-grained relabeling (MVV)": "https://raw.githubusercontent.com/nec-labs-ma/relabeling/main/LICENSE.txt",
    "Belgium Traffic Sign Dataset (BelgiumTS/BTSC)": "https://btsd.ethz.ch/shareddata/",
    "Tsinghua-Tencent 100K (TT100K)": "https://cg.cs.tsinghua.edu.cn/ctsdb/",
}

MANUAL_ASSESSMENTS = {
    "img_0579.jpg": (
        "no_left_turn",
        "yes",
        "Clear circular no-left-turn sign; requires a new unseen review.",
    ),
    "img_1021.jpg": (
        "no_left_turn",
        "yes",
        "Distant but visible no-left-turn sign; requires a new unseen review.",
    ),
    "img_0786.jpg": (
        "no_parking",
        "yes",
        "Visible no-parking sign; same scene/source dependency as img_0812.jpg.",
    ),
    "img_0812.jpg": (
        "no_parking",
        "yes",
        "Visible no-parking sign; same scene/source dependency as img_0786.jpg.",
    ),
    "img_0816.jpg": (
        "no_parking",
        "yes",
        "Visible no-parking sign in a different roadside scene.",
    ),
    "img_0970.jpg": (
        "no_parking",
        "yes",
        "Police no-parking board; visually valid but style differs from standard fixed signs.",
    ),
    "img_1016.jpg": (
        "no_parking",
        "yes",
        "Visible no-parking regulatory sign in a distinct scene.",
    ),
    "img_1046.jpg": (
        "no_parking",
        "no",
        "Visual evidence is commercial 'store/buy stop' signage, not a road no-parking sign.",
    ),
    "img_1024.jpg": (
        "maximum_speed_limit_50_km_h",
        "no",
        "The photograph visibly shows a 60 km/h sign; the VQA 50 km/h annotation is wrong.",
    ),
    "img_1035.jpg": (
        "maximum_speed_limit_50_km_h",
        "no",
        "The photograph shows commercial food/shop signage, not a speed-limit sign.",
    ),
}

DATASET_A_TEMPLATES = (
    ("15", "58.png", "no_left_turn"),
    ("21", "64.png", "no_parking"),
    ("44", "87.png", "roundabout_ahead"),
)


def main() -> int:
    """Generate and validate the source-discovery package without acquisition."""
    try:
        recoveries = _build_recoveries()
        counts = eligible_counts_by_class(recoveries)
        _validate_measured_recoveries(recoveries, counts)
        availability = _availability_rows(counts)
        candidate_rows = _candidate_rows(availability)
        license_rows = _license_rows()
        recommendation_rows = _recommendation_rows(counts)
        OUTPUT.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(
            prefix="source-discovery-", dir=OUTPUT
        ) as temp:
            stage = Path(temp)
            staged_pdf = stage / PDF_PATH.name
            staged_xlsx = stage / XLSX_PATH.name
            staged_csv = stage / CSV_PATH.name
            _write_csv(staged_csv, availability)
            _write_workbook(
                staged_xlsx,
                {
                    "Candidate Classes": candidate_rows,
                    "Dataset Sources": list(DATASET_SOURCES),
                    "License Assessment": license_rows,
                    "Local Recoveries": recoveries,
                    "Availability": availability,
                    "Recommendation": recommendation_rows,
                },
            )
            _write_pdf(
                staged_pdf,
                candidate_rows,
                license_rows,
                recoveries,
                availability,
                recommendation_rows,
            )
            _validate_outputs(staged_pdf, staged_xlsx, staged_csv, recoveries)
            os.replace(staged_pdf, PDF_PATH)
            os.replace(staged_xlsx, XLSX_PATH)
            os.replace(staged_csv, CSV_PATH)
        output = {
            "artifacts": {
                path.name: {"path": str(path), "sha256": _sha256(path)}
                for path in (PDF_PATH, XLSX_PATH, CSV_PATH)
            },
            "pdf_pages": len(PdfReader(str(PDF_PATH)).pages),
            "local_recovery_rows": len(recoveries),
            "eligible_counts": counts,
            "frozen_split_overlaps": sum(
                row["overlap_with_frozen_v2_splits"] == "yes" for row in recoveries
            ),
            "download_performed": False,
            "training_evaluation_or_calibration_performed": False,
        }
        print(json.dumps(output, ensure_ascii=False, indent=2))
        return 0
    except (OSError, ValueError, SourceDiscoveryError) as error:
        print(f"FATAL: {error}", file=sys.stderr)
        return 2


def _build_recoveries() -> list[dict[str, str]]:
    inventory = _read_csv(OUTPUT.parent / "dataset_b_audit/image_inventory.csv")
    image_hashes = {row["image_name"]: row["sha256"] for row in inventory}
    near_pairs = _read_csv(OUTPUT.parent / "dataset_b_audit/near_duplicate_pairs.csv")
    groups = build_perceptual_groups(image_hashes, near_pairs)
    split_rows: list[dict[str, str]] = []
    for name in ("v2_train.csv", "v2_validation.csv", "v2_test.csv"):
        split_rows.extend(_read_csv(OUTPUT.parent / f"manifests/{name}"))
    split_evidence = make_split_evidence(split_rows, image_hashes)
    excluded = {
        row["image_id"]: row
        for row in _read_csv(OUTPUT.parent / "dataset_b_audit/excluded_candidates.csv")
    }
    review = _read_csv(OUTPUT.parent / "v2_review/dataset_b_v2_review.csv")
    rows: list[dict[str, str]] = []
    for source_id, (new_label, visual_eligible, reason) in MANUAL_ASSESSMENTS.items():
        source = excluded[source_id]
        rows.append(
            _dataset_b_recovery_row(
                source_id=source_id,
                original_label=source.get("proposed_label", ""),
                original_outcome=(f"excluded_before_v2_review:{source['reason']}"),
                new_label=new_label,
                visual_eligible=visual_eligible,
                reason=reason,
                groups=groups,
                image_hashes=image_hashes,
                split_evidence=split_evidence,
            )
        )
    for source in review:
        if source["proposed_class"] != "stop":
            continue
        rows.append(
            _dataset_b_recovery_row(
                source_id=source["source_image_id"],
                original_label="stop",
                original_outcome=f"V2_{source['review_status']}",
                new_label="bus_stop",
                visual_eligible="yes",
                reason=(
                    "Existing contact-sheet audit shows a bus-stop or bus-bay sign, "
                    "not a STOP sign. This is an alternative-class candidate only."
                ),
                groups=groups,
                image_hashes=image_hashes,
                split_evidence=split_evidence,
            )
        )
    for class_id, filename, label in DATASET_A_TEMPLATES:
        path = DATASET_A_IMAGE_ROOT / class_id / filename
        rows.append(
            {
                "source_dataset": "Dataset A template/augmentation corpus",
                "image_source_id": f"dataset_a_class_{class_id}_template_{path.stem}",
                "image_path": path.relative_to(PROJECT_ROOT).as_posix(),
                "original_proposed_label": label,
                "original_v2_review_outcome": "not_in_v2; Dataset A audit found one source template per populated class",
                "newly_proposed_unseen_label": label,
                "reason": "Only one 32x32 isolated sign template exists; 200 other files are dependent augmentations, not independent real photographs.",
                "perceptual_group": f"dataset_a_class_{class_id}_template_{path.stem}",
                "conservative_dependency_group": f"dataset_a_class_{class_id}_template_{path.stem}",
                "overlap_with_frozen_v2_splits": "no",
                "overlap_reason": "different source corpus; not present in V2 manifests",
                "eligible_for_unseen_review": "no",
                "new_unseen_review_status": "not_queued",
                "license_blocker": "source archive rights not documented in repository audit",
            }
        )
    return sorted(
        rows,
        key=lambda row: (
            row["newly_proposed_unseen_label"],
            row["image_source_id"],
        ),
    )


def _dataset_b_recovery_row(
    *,
    source_id: str,
    original_label: str,
    original_outcome: str,
    new_label: str,
    visual_eligible: str,
    reason: str,
    groups: dict[str, str],
    image_hashes: dict[str, str],
    split_evidence: Any,
) -> dict[str, str]:
    group_id = groups[source_id]
    overlap, overlap_reason = frozen_split_overlap(
        source_id=source_id,
        perceptual_group_id=group_id,
        sha256=image_hashes[source_id],
        evidence=split_evidence,
    )
    dependency_group = group_id
    if source_id in {"img_0786.jpg", "img_0812.jpg"}:
        dependency_group = "manual_source_group_no_parking_roadside_01"
    eligible = visual_eligible == "yes" and not overlap
    return {
        "source_dataset": "Indian Traffic VQA Dataset B",
        "image_source_id": source_id,
        "image_path": f"data/raw/indian_traffic_vqa/traffic512final/{source_id}",
        "original_proposed_label": original_label,
        "original_v2_review_outcome": original_outcome,
        "newly_proposed_unseen_label": new_label,
        "reason": reason,
        "perceptual_group": group_id,
        "conservative_dependency_group": dependency_group,
        "overlap_with_frozen_v2_splits": "yes" if overlap else "no",
        "overlap_reason": overlap_reason,
        "eligible_for_unseen_review": "yes" if eligible else "no",
        "new_unseen_review_status": "pending" if eligible else "not_queued",
        "license_blocker": "Confirm that the Hugging Face research/distribution terms apply to the identical Zenodo image files before experiment use or repository redistribution.",
    }


def _validate_measured_recoveries(
    rows: list[dict[str, str]], counts: dict[str, dict[str, int]]
) -> None:
    expected = {
        "bus_stop": {"eligible_images": 36, "independent_groups": 31},
        "no_left_turn": {"eligible_images": 2, "independent_groups": 2},
        "no_parking": {"eligible_images": 5, "independent_groups": 4},
    }
    if len(rows) != 49 or counts != expected:
        raise SourceDiscoveryError(
            f"Unexpected recovery result: rows={len(rows)}, counts={counts}"
        )
    if any(row["overlap_with_frozen_v2_splits"] == "yes" for row in rows):
        raise SourceDiscoveryError("A recovery candidate overlaps a frozen V2 split")
    if any(row["new_unseen_review_status"] == "approved" for row in rows):
        raise SourceDiscoveryError("Recovery workflow must never auto-approve images")


def _availability_rows(counts: dict[str, dict[str, int]]) -> list[dict[str, Any]]:
    routes = {
        "stop": (
            "Mapillary India metadata-first query; IRSDB request; GTSRB cross-domain fallback",
            "B",
            "No local STOP images. Obtain at least 15 independent groups; target 30-50.",
            "Keep class; distinctive and well supported internationally.",
        ),
        "no_left_turn": (
            "Mapillary India metadata-first query; IRSDB request; LISA fallback",
            "B",
            "2 local pending candidates / 2 groups; acquire at least 13 more independent groups.",
            "Keep class; direct confusion test against V2 no_right_turn.",
        ),
        "maximum_speed_limit_50_km_h": (
            "Mapillary India metadata-first query; GTSRB cross-domain fallback",
            "B",
            "Both local 50 annotations failed visual audit; acquire at least 15 independent groups.",
            "Keep class if Indian Mapillary/IRSDB examples are found; GTSRB alone is cross-domain.",
        ),
        "no_parking": (
            "Mapillary India metadata-first query; IRSDB request; LISA/MVV fallback",
            "B",
            "5 local pending candidates / 4 conservative groups; acquire at least 11 more independent groups.",
            "Keep class; useful regulatory-family confusion test.",
        ),
        "bus_stop": (
            "Local Dataset B candidates after curator licence confirmation",
            "A after human review and licence confirmation",
            "36 local pending candidates / 31 conservative groups; do not approve automatically.",
            "Locked replacement for roundabout_ahead; retain all source and dependency evidence.",
        ),
    }
    rows: list[dict[str, Any]] = []
    for class_name in TARGET_CLASSES:
        local = counts.get(class_name, {"eligible_images": 0, "independent_groups": 0})
        best_source, route, requirement, replacement = routes[class_name]
        rows.append(
            {
                "class_name": class_name,
                "local_recoverable_images_pending_review": local["eligible_images"],
                "local_conservative_independent_groups": local["independent_groups"],
                "minimum_required_groups": 15,
                "target_photos": "30-50",
                "best_source_route": best_source,
                "recommendation_code": route,
                "external_acquisition_requirement": requirement,
                "replacement_assessment": replacement,
                "enough_data_now": "no",
            }
        )
    return rows


def _candidate_rows(availability: list[dict[str, Any]]) -> list[dict[str, Any]]:
    roles = {
        "stop": "distinctive",
        "no_left_turn": "highly confusable with no_right_turn",
        "maximum_speed_limit_50_km_h": "highly confusable with V2 speed limits",
        "no_parking": "moderately similar regulatory sign",
        "bus_stop": "distinctive public-transport sign",
    }
    return [
        {
            "class_name": row["class_name"],
            "experimental_role": roles[str(row["class_name"])],
            "outside_v2_17_classes": "yes",
            "local_pending_images": row["local_recoverable_images_pending_review"],
            "local_groups": row["local_conservative_independent_groups"],
            "best_source_route": row["best_source_route"],
            "status": "blocked_pending_licensed_acquisition_and_human_review",
        }
        for row in availability
    ]


def _license_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for source in DATASET_SOURCES:
        name = source["source_name"]
        redistributable = source["repository_redistribution"]
        decision = (
            "metadata/permission inquiry only"
            if redistributable.startswith("no")
            or redistributable.startswith("conditional-no")
            else "candidate after terms snapshot and attribution plan"
        )
        rows.append(
            {
                "source_name": name,
                "rights_evidence_url": LICENSE_URLS[name],
                "stated_terms": source["license_research_terms"],
                "repository_redistribution": redistributable,
                "legal_confidence": source["legal_confidence"],
                "current_download_decision": decision,
                "required_action": _required_legal_action(name),
            }
        )
    return rows


def _required_legal_action(name: str) -> str:
    actions = {
        "Indian Traffic VQA Dataset (local Dataset B)": "Ask curator to bind one explicit licence to DOI 10.5281/zenodo.17300841 and confirm repository redistribution/crops.",
        "IRSDBv1.0 Fully Annotated Indian Traffic Signs Database": "Request access, class manifest, and written permission for private-repository storage and derived crops.",
        "Mapillary platform imagery and traffic-sign map features": "Run metadata-only India query after normal authentication; snapshot current terms and retain per-image author/profile attribution.",
        "Mapillary Traffic Sign Dataset (MTSD)": "Use only internally after accepting research licence; do not copy images or derivatives into repository/report.",
        "German Traffic Sign Recognition Benchmark (GTSRB)": "Inspect archive readme and record citation before acquiring only stop and speed-50 tracks.",
        "LISA Traffic Signs Dataset": "Obtain and review the academic licence agreement before download.",
        "DATS_2022": "Inspect annotations remotely/metadata-first for target signs; preserve CC BY attribution and video-source groups.",
        "Mapillary Vistas Validation fine-grained relabeling (MVV)": "Confirm that dataset-image redistribution, not only software use, is covered by the research licence and Mapillary terms.",
        "Belgium Traffic Sign Dataset (BelgiumTS/BTSC)": "Ask maintainer for explicit research reuse and redistribution permission.",
        "Tsinghua-Tencent 100K (TT100K)": "Obtain explicit dataset licence/permission before download or use.",
    }
    return actions[name]


def _recommendation_rows(counts: dict[str, dict[str, int]]) -> list[dict[str, Any]]:
    availability = _availability_rows(counts)
    rows = [
        {
            "class_name": row["class_name"],
            "recommendation": row["recommendation_code"],
            "recommended_action": row["best_source_route"],
            "why": row["external_acquisition_requirement"],
            "protocol_change": "none",
        }
        for row in availability
    ]
    return rows


def _write_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        sheet = workbook.create_sheet(name)
        if not rows:
            raise SourceDiscoveryError(f"Workbook sheet would be empty: {name}")
        columns = list(rows[0])
        sheet.append(columns)
        for row in rows:
            sheet.append([row.get(column, "") for column in columns])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 30
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_index in range(1, sheet.max_column + 1):
            width = max(
                len(str(sheet.cell(row_index, column_index).value or ""))
                for row_index in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(width + 2, 12), 55
            )
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        raise SourceDiscoveryError("Summary CSV cannot be empty")
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def _write_pdf(
    path: Path,
    candidates: list[dict[str, Any]],
    licenses: list[dict[str, str]],
    recoveries: list[dict[str, str]],
    availability: list[dict[str, Any]],
    recommendations: list[dict[str, Any]],
) -> None:
    styles = getSampleStyleSheet()
    small = ParagraphStyle(
        "Small",
        parent=styles["BodyText"],
        fontSize=7,
        leading=9,
        spaceAfter=1,
    )
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Open-set unseen-class source discovery report",
        author="Adaptive Indian Road Sign Recognition project",
    )
    story: list[Any] = [
        Paragraph("Open-Set Unseen-Class Source Discovery Report", styles["Title"]),
        Paragraph(
            "Evidence date: 2026-09-01. Source discovery and local re-audit only; no dataset download, prototype creation, calibration, training, or evaluation was performed.",
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
        _pdf_table(
            [["Finding", "Measured result"]]
            + [
                [
                    "Frozen Baseline V2",
                    "17 classes; manifests read only; zero candidate source/group/hash overlaps",
                ],
                [
                    "Target-class local recovery",
                    "bus_stop: 36 images / 31 groups; no_left_turn: 2 / 2; no_parking: 5 / 4; stop and speed-50: 0",
                ],
                [
                    "Locked protocol change",
                    "bus_stop replaces roundabout_ahead; all 36 local candidates remain pending a new label review",
                ],
                [
                    "Minimum protocol",
                    "15 independent approved groups per class; target 30-50 photos",
                ],
                [
                    "Current experiment status",
                    "Blocked: no target class has enough approved, legally cleared data",
                ],
            ],
            [55 * mm, 205 * mm],
            9,
        ),
        PageBreak(),
        Paragraph("Candidate classes and availability", styles["Heading1"]),
        _rows_table(
            candidates,
            (
                "class_name",
                "experimental_role",
                "local_pending_images",
                "local_groups",
                "best_source_route",
                "status",
            ),
            [40 * mm, 48 * mm, 22 * mm, 20 * mm, 100 * mm, 45 * mm],
            7,
        ),
        Spacer(1, 4 * mm),
        Paragraph("Recommended acquisition route", styles["Heading2"]),
        _rows_table(
            recommendations,
            (
                "class_name",
                "recommendation",
                "recommended_action",
                "why",
                "protocol_change",
            ),
            [42 * mm, 35 * mm, 80 * mm, 75 * mm, 50 * mm],
            7,
        ),
        PageBreak(),
        Paragraph("Candidate dataset sources", styles["Heading1"]),
        _rows_table(
            list(DATASET_SOURCES),
            (
                "source_name",
                "geographic_domain",
                "relevant_classes",
                "documented_size",
                "real_photographs",
                "scientific_confidence",
            ),
            [48 * mm, 29 * mm, 80 * mm, 48 * mm, 25 * mm, 52 * mm],
            6,
        ),
        PageBreak(),
        Paragraph("Licence and redistribution assessment", styles["Heading1"]),
        _rows_table(
            licenses,
            (
                "source_name",
                "stated_terms",
                "repository_redistribution",
                "legal_confidence",
                "current_download_decision",
                "required_action",
            ),
            [45 * mm, 65 * mm, 50 * mm, 30 * mm, 40 * mm, 52 * mm],
            6,
        ),
        Spacer(1, 3 * mm),
        Paragraph(
            "Public accessibility is not treated as a licence. MTSD, LISA, IRSDB, BelgiumTS, TT100K, and MVV remain restricted or ambiguous for repository redistribution. Mapillary platform imagery requires current-terms compliance and per-image attribution. GTSRB is an international fallback, not evidence of Indian-domain performance.",
            small,
        ),
        PageBreak(),
        Paragraph("Local recovery analysis", styles["Heading1"]),
        _pdf_table(
            [
                ["Proposed label", "Eligible images", "Conservative groups", "Status"],
                [
                    "no_left_turn",
                    "2",
                    "2",
                    "new human review pending; licence confirmation pending",
                ],
                [
                    "no_parking",
                    "5",
                    "4",
                    "new human review pending; img_0786/img_0812 grouped as one scene",
                ],
                [
                    "maximum_speed_limit_50_km_h",
                    "0",
                    "0",
                    "two annotations visually disproved",
                ],
                ["stop", "0", "0", "36 V2 rows are bus-stop signs, not STOP signs"],
                [
                    "bus_stop",
                    "36",
                    "31",
                    "new review required; no old decision reused",
                ],
            ],
            [65 * mm, 35 * mm, 38 * mm, 140 * mm],
            8,
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "All 49 recovery-table rows have zero source-image, perceptual-group, and exact-SHA overlap with V2 train/validation/test. The XLSX Local Recoveries sheet preserves every source ID, original outcome, new proposal, group, overlap result, and pending/not-queued status.",
            styles["BodyText"],
        ),
        PageBreak(),
        Paragraph(
            "Visual evidence: ten target-class Dataset B candidates", styles["Heading1"]
        ),
        Paragraph(
            "Images are shown only inside this report. Captions record source ID, new proposal, and eligibility; none is approved.",
            small,
        ),
        _recovery_visual_table(recoveries, small, list(MANUAL_ASSESSMENTS)[:6]),
        PageBreak(),
        Paragraph(
            "Visual evidence: target-class candidates (continued)", styles["Heading1"]
        ),
        _recovery_visual_table(recoveries, small, list(MANUAL_ASSESSMENTS)[6:]),
        PageBreak(),
        Paragraph(
            "Visual evidence: rejected STOP rows are bus-stop signs (1/3)",
            styles["Heading1"],
        ),
        _full_page_image(OUTPUT.parent / "v2_review/contact_sheets/stop_01.jpg"),
        PageBreak(),
        Paragraph(
            "Visual evidence: rejected STOP rows are bus-stop signs (2/3)",
            styles["Heading1"],
        ),
        _full_page_image(OUTPUT.parent / "v2_review/contact_sheets/stop_02.jpg"),
        PageBreak(),
        Paragraph(
            "Visual evidence: rejected STOP rows are bus-stop signs (3/3)",
            styles["Heading1"],
        ),
        _full_page_image(OUTPUT.parent / "v2_review/contact_sheets/stop_03.jpg"),
        PageBreak(),
        Paragraph("Dataset A template-only evidence", styles["Heading1"]),
        Paragraph(
            "These isolated 32x32 templates are not independent real photographs. Each class has one original template and 200 dependent augmentations.",
            styles["BodyText"],
        ),
        _dataset_a_visual_table(small),
        PageBreak(),
        Paragraph("Blockers and exact next action", styles["Heading1"]),
        _pdf_table(
            [
                ["Priority", "Action", "Why"],
                [
                    "1",
                    "Ask the Indian Traffic VQA curator to confirm one licence for the exact Zenodo files and repository/crop redistribution.",
                    "Local recoveries cannot enter the experiment on an ambiguous Rights record.",
                ],
                [
                    "2",
                    "Request IRSDBv1.0 access, its 49-class manifest, per-class track counts, and explicit private-repository/derived-crop permission.",
                    "Best India-specific source with physical-sign tracks.",
                ],
                [
                    "3",
                    "After normal Mapillary authentication, run a metadata-only India query for the five exact sign classes; record image IDs, sequences, contributors, and current terms before downloading pixels.",
                    "Best scalable India-domain discovery route; prevents unclear or duplicate acquisition.",
                ],
                [
                    "4",
                    "If India sources cannot supply 15 groups, acquire GTSRB stop/speed-50 tracks as explicitly cross-domain calibration/query data, not India-domain evidence.",
                    "Clearer reuse statement and strong real-photo grouping.",
                ],
                [
                    "5",
                    "Use the locked bus_stop replacement protocol and complete new human review of the 36 candidates.",
                    "The prior STOP rejection is not evidence of bus_stop approval; licence confirmation is still required.",
                ],
            ],
            [20 * mm, 165 * mm, 95 * mm],
            8,
        ),
        Spacer(1, 4 * mm),
        Paragraph("Evidence URLs", styles["Heading2"]),
    ]
    for source in DATASET_SOURCES:
        story.append(
            Paragraph(
                f"<b>{_escape(source['source_name'])}</b>: {_escape(source['official_url'])}",
                small,
            )
        )
    document.build(story)


def _recovery_visual_table(
    recoveries: list[dict[str, str]],
    style: ParagraphStyle,
    ids: list[str],
) -> Table:
    by_id = {row["image_source_id"]: row for row in recoveries}
    cells: list[list[Any]] = []
    for start in range(0, len(ids), 2):
        row_cells: list[Any] = []
        for source_id in ids[start : start + 2]:
            row = by_id[source_id]
            row_cells.append(
                [
                    _scaled_image(VQA_IMAGE_ROOT / source_id, 108 * mm, 37 * mm),
                    Paragraph(
                        f"{source_id} | {row['newly_proposed_unseen_label']} | eligible: {row['eligible_for_unseen_review']}<br/>{_escape(row['reason'])}",
                        style,
                    ),
                ]
            )
        cells.append(row_cells)
    table = Table(
        cells,
        colWidths=[137 * mm, 137 * mm],
        rowHeights=[52 * mm] * len(cells),
    )
    table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    return table


def _dataset_a_visual_table(style: ParagraphStyle) -> Table:
    cells: list[Any] = []
    for class_id, filename, label in DATASET_A_TEMPLATES:
        path = DATASET_A_IMAGE_ROOT / class_id / filename
        cells.append(
            [
                _scaled_image(path, 35 * mm, 35 * mm),
                Paragraph(
                    f"Class {class_id}: {label}<br/>{filename}<br/>Not eligible",
                    style,
                ),
            ]
        )
    table = Table([cells], colWidths=[90 * mm] * 3)
    table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    return table


def _scaled_image(path: Path, maximum_width: float, maximum_height: float) -> Image:
    if not path.is_file():
        raise SourceDiscoveryError(f"Visual evidence image is missing: {path}")
    with PilImage.open(path) as source:
        width, height = source.size
    scale = min(maximum_width / width, maximum_height / height)
    return Image(str(path), width=width * scale, height=height * scale)


def _full_page_image(path: Path) -> Image:
    return _scaled_image(path, 268 * mm, 170 * mm)


def _rows_table(
    rows: list[dict[str, Any]],
    columns: tuple[str, ...],
    widths: list[float],
    font_size: int,
) -> Table:
    data: list[list[Any]] = [[_header(column) for column in columns]]
    for row in rows:
        data.append([str(row.get(column, "")) for column in columns])
    return _pdf_table(data, widths, font_size)


def _pdf_table(data: list[list[Any]], widths: list[float], font_size: int) -> Table:
    body_style = ParagraphStyle(
        f"TableBody{font_size}",
        fontName="Helvetica",
        fontSize=font_size,
        leading=font_size + 2,
        spaceAfter=0,
    )
    header_style = ParagraphStyle(
        f"TableHeader{font_size}",
        parent=body_style,
        fontName="Helvetica-Bold",
        textColor=colors.white,
    )
    formatted: list[list[Any]] = []
    for row_index, row in enumerate(data):
        formatted.append(
            [
                (
                    Paragraph(
                        _escape(str(cell)),
                        header_style if row_index == 0 else body_style,
                    )
                    if isinstance(cell, (str, int, float))
                    else cell
                )
                for cell in row
            ]
        )
    table = Table(formatted, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("LEADING", (0, 0), (-1, -1), font_size + 2),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#EDF3F8")],
                ),
            ]
        )
    )
    return table


def _validate_outputs(
    pdf: Path, xlsx: Path, csv_path: Path, recoveries: list[dict[str, str]]
) -> None:
    page_count = len(PdfReader(str(pdf)).pages)
    if page_count < 10:
        raise SourceDiscoveryError(f"PDF is unexpectedly short: {page_count} pages")
    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    expected_sheets = [
        "Candidate Classes",
        "Dataset Sources",
        "License Assessment",
        "Local Recoveries",
        "Availability",
        "Recommendation",
    ]
    if workbook.sheetnames != expected_sheets:
        raise SourceDiscoveryError(f"Unexpected workbook sheets: {workbook.sheetnames}")
    if workbook["Local Recoveries"].max_row - 1 != len(recoveries):
        raise SourceDiscoveryError("Local Recoveries workbook row count mismatch")
    workbook.close()
    summary_rows = _read_csv(csv_path)
    if len(summary_rows) != 5 or {row["class_name"] for row in summary_rows} != set(
        TARGET_CLASSES
    ):
        raise SourceDiscoveryError("Summary CSV must contain exactly five target rows")


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise SourceDiscoveryError(f"CSV has no header: {path}")
        return [dict(row) for row in reader]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _escape(value: str) -> str:
    return value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _header(value: str) -> str:
    return value.replace("_", " ").title()


if __name__ == "__main__":
    raise SystemExit(main())
