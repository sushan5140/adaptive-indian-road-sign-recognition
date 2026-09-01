"""Package measured Baseline V2 artifacts as reviewable PDF and workbooks."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "outputs" / "v2_results"
RUN_ID = json.loads((RESULTS / "run_artifacts.json").read_text(encoding="utf-8"))[
    "run_id"
]
RUN = ROOT / "outputs" / "runs" / RUN_ID
EMBEDDING = ROOT / "outputs" / "embedding_analysis" / "v2_embedding_summary.json"


def _json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _format_workbook(workbook: Workbook) -> None:
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for column in range(1, sheet.max_column + 1):
            width = max(
                len(str(sheet.cell(row, column).value or ""))
                for row in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[get_column_letter(column)].width = min(
                max(width + 2, 12), 42
            )


def _write_sheet(path: Path, title: str, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([row[key] for key in headers])
    _format_workbook(workbook)
    workbook.save(path)


def _table(
    data: list[list[Any]], widths: list[float] | None = None, *, font_size: int = 7
) -> Table:
    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#EDF3F8")],
                ),
            ]
        )
    )
    return table


def main() -> None:
    """Create the three XLSX files, comparison note, and PDF report."""
    history = _json(RUN / "history.json")
    metrics = _json(RESULTS / "metrics.json")
    summary = _json(RESULTS / "v2_experiment_summary.json")
    environment = _json(RUN / "environment.json")
    weights = _json(RESULTS / "class_weights.json")
    preflight = _json(RESULTS / "pretraining_verification.json")
    embedding = _json(EMBEDDING)
    per_class = _csv(RESULTS / "per_class_metrics.csv")
    confusion_rows = _csv(RESULTS / "confusion_matrix.csv")
    for row in history:
        row["epoch"] = int(row["epoch"]) + 1
    _write_sheet(RESULTS / "v2_training_history.xlsx", "Training History", history)
    matrix = metrics["confusion_matrix"]
    for index, row in enumerate(per_class):
        row["correct"] = matrix[index][index]
        row["total"] = row["support"]
    _write_sheet(RESULTS / "v2_per_class_metrics.xlsx", "Per-Class Metrics", per_class)
    _write_sheet(
        RESULTS / "v2_confusion_matrix.xlsx", "Confusion Matrix", confusion_rows
    )

    comparison_dir = ROOT / "outputs" / "comparisons"
    comparison_dir.mkdir(parents=True, exist_ok=True)
    comparison = f"""# Baseline V1 vs Baseline V2

| Item | Baseline V1 | Baseline V2 |
|---|---:|---:|
| Data design | Dataset A template/augmentation-derived training; external real-world evaluation | Real-world, manually reviewed, group-safe splits |
| Classes | 5 | 17 |
| Test accuracy | 14.53% (external Dataset B) | {metrics['top1_accuracy']:.2%} |
| Test macro F1 | 6.54% | {metrics['macro_f1']:.2%} |

V1 showed a severe synthetic/template-to-real-world domain gap. V2 uses 17 real-world classes and a validation-selected checkpoint, with the locked test split evaluated once after training. These are not directly equivalent benchmark scores because their datasets, class sets, and experimental designs differ.
"""
    (comparison_dir / "v1_vs_v2.md").write_text(comparison, encoding="utf-8")

    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph("Baseline V2 Training and Evaluation Report", styles["Title"]),
        Paragraph(
            "Adaptive Indian Road Sign Recognition Using Open-Set and Few-Shot Learning",
            styles["Heading2"],
        ),
        Spacer(1, 5 * mm),
        Paragraph(
            "Measured closed-set Baseline V2 run. No post-test tuning or retraining was performed.",
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
        _table(
            [
                ["Item", "Value"],
                ["Run ID", RUN_ID],
                ["Python", environment["python_version"]],
                ["PyTorch", environment["torch_version"]],
                ["timm", environment["timm_version"]],
                ["Device", environment["selected_device"]],
                ["Deterministic", str(environment["deterministic_algorithms"])],
                ["Train / validation / test", "287 / 62 / 63"],
                ["Classes", "17"],
                ["Best epoch", summary["best_epoch_one_based"]],
                [
                    "Best validation macro F1",
                    f"{summary['best_validation_macro_f1']:.6f}",
                ],
                ["Test accuracy", f"{metrics['top1_accuracy']:.6f}"],
                ["Test macro F1", f"{metrics['macro_f1']:.6f}"],
            ],
            [55 * mm, 190 * mm],
            font_size=9,
        ),
        PageBreak(),
        Paragraph("Pre-training verification and class weights", styles["Heading1"]),
        Paragraph(
            f"Review: 530 rows; 412 approved; 118 rejected; zero pending. Leakage checks passed for review IDs, source images, and perceptual groups. Excluded classes one_way, stop, and y_junction were absent. Locked file hashes are recorded in pretraining_verification.json.",
            styles["BodyText"],
        ),
        Spacer(1, 3 * mm),
        Paragraph(
            f"Weight formula: {weights['formula']}; source: train split only.",
            styles["BodyText"],
        ),
        Spacer(1, 2 * mm),
        _table(
            [
                ["Class", "Train count", "Weight"],
                *[
                    [
                        label,
                        weights["counts"][label],
                        f"{weights['weights'][label]:.6f}",
                    ]
                    for label in weights["counts"]
                ],
            ],
            [80 * mm, 35 * mm, 40 * mm],
            font_size=7,
        ),
        PageBreak(),
        Paragraph("30-epoch training history", styles["Heading1"]),
        _table(
            [
                [
                    "Epoch",
                    "Train loss",
                    "Train acc.",
                    "Val. loss",
                    "Val. acc.",
                    "Val. macro F1",
                    "LR",
                    "Seconds",
                ],
                *[
                    [
                        row["epoch"],
                        f"{row['train_loss']:.4f}",
                        f"{row['train_accuracy']:.4f}",
                        f"{row['val_loss']:.4f}",
                        f"{row['val_accuracy']:.4f}",
                        f"{row['val_macro_f1']:.4f}",
                        f"{row['learning_rate']:.8f}",
                        f"{row['elapsed_seconds']:.1f}",
                    ]
                    for row in history
                ],
            ],
            [14 * mm, 25 * mm, 22 * mm, 25 * mm, 22 * mm, 28 * mm, 26 * mm, 20 * mm],
            font_size=6,
        ),
        PageBreak(),
        Paragraph("Locked test results", styles["Heading1"]),
        _table(
            [
                ["Metric", "Value"],
                ["Top-1 accuracy", f"{metrics['top1_accuracy']:.6f}"],
                ["Macro precision", f"{metrics['macro_precision']:.6f}"],
                ["Macro recall", f"{metrics['macro_recall']:.6f}"],
                ["Macro F1", f"{metrics['macro_f1']:.6f}"],
                ["Weighted precision", f"{metrics['weighted_precision']:.6f}"],
                ["Weighted recall", f"{metrics['weighted_recall']:.6f}"],
                ["Weighted F1", f"{metrics['weighted_f1']:.6f}"],
                ["Correct / total", f"{summary['correct']} / 63"],
            ],
            [60 * mm, 45 * mm],
            font_size=8,
        ),
        Spacer(1, 4 * mm),
        _table(
            [
                ["Class", "Precision", "Recall", "F1", "Correct/total"],
                *[
                    [
                        row["label"],
                        f"{float(row['precision']):.3f}",
                        f"{float(row['recall']):.3f}",
                        f"{float(row['f1']):.3f}",
                        f"{row['correct']}/{row['total']}",
                    ]
                    for row in per_class
                ],
            ],
            [78 * mm, 28 * mm, 28 * mm, 28 * mm, 30 * mm],
            font_size=7,
        ),
        PageBreak(),
        Paragraph("Confusions, embeddings, and limitations", styles["Heading1"]),
        Paragraph("Most common non-diagonal confusion pairs:", styles["BodyText"]),
        _table(
            [
                ["True class", "Predicted class", "Count"],
                *[
                    [item["true_label"], item["predicted_label"], item["count"]]
                    for item in summary["most_common_confusions"]
                ],
            ],
            [80 * mm, 80 * mm, 25 * mm],
            font_size=7,
        ),
        Spacer(1, 4 * mm),
        _table(
            [
                ["Embedding diagnostic", "Value"],
                ["Dimension", embedding["embedding_dimension"]],
                [
                    "Mean within-class cosine",
                    f"{embedding['mean_within_class_cosine_similarity']:.6f}",
                ],
                [
                    "Mean between-class cosine",
                    f"{embedding['mean_between_class_cosine_similarity']:.6f}",
                ],
                [
                    "Nearest-centroid closed-set accuracy",
                    f"{embedding['nearest_centroid_closed_set_accuracy']:.6f}",
                ],
            ],
            [85 * mm, 80 * mm],
            font_size=8,
        ),
        Spacer(1, 4 * mm),
        Paragraph("Warnings and limitations", styles["Heading2"]),
        Paragraph(
            "no_right_turn and pass_either_side each have only one test image, so their class metrics are unstable. Several other class supports are also small. The nearest-centroid value is a resubstitution diagnostic because centroids use the same test embeddings, including each query; it is not an unbiased model score. Softmax confidence is closed-set confidence, not an unknown-sign score. No unknown/OOD threshold was implemented. V1 and V2 are descriptive, not directly comparable benchmarks.",
            styles["BodyText"],
        ),
    ]
    pdf = SimpleDocTemplate(
        str(RESULTS / "v2_training_and_evaluation_report.pdf"),
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    pdf.build(story)
    assert preflight["passed"] and len(history) == 30 and len(per_class) == 17


if __name__ == "__main__":
    main()
