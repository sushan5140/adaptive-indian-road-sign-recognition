"""Freeze no-parking metadata and expand India-only Mapillary discovery."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
import shutil
import sys
import tempfile
from collections import Counter, defaultdict
from collections.abc import Mapping, Sequence
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from data.mapillary_acquisition import (  # noqa: E402
    EXPANDED_TARGET_CLASSES,
    build_no_parking_acquisition_plan,
    merge_expanded_candidates,
)
from data.mapillary_metadata import (  # noqa: E402
    MapillaryMetadataError,
    map_exact_taxonomy_label,
)
from scripts.discover_mapillary_india_metadata import (  # noqa: E402
    CANDIDATE_COLUMNS,
    CITY_CENTRES,
    MapillaryClient,
    _build_candidate_rows,
    _coordinates,
    _deduplicate_candidate_images,
    _feature_image_associations,
    _fetch_image_metadata,
)

OUTPUT_DIR = PROJECT_ROOT / "outputs/open_set_acquisition"
R07_CSV = OUTPUT_DIR / "mapillary_india_metadata_20260901_r07.csv"
PLAN_PDF = OUTPUT_DIR / "no_parking_pixel_acquisition_plan_20260901_r02.pdf"
PLAN_XLSX = OUTPUT_DIR / "no_parking_pixel_acquisition_plan_20260901_r02.xlsx"
PLAN_CSV = OUTPUT_DIR / "no_parking_pixel_acquisition_plan_20260901_r02.csv"
EXPANDED_PDF = OUTPUT_DIR / "mapillary_india_expanded_metadata_report_20260901_r02.pdf"
EXPANDED_XLSX = OUTPUT_DIR / "mapillary_india_expanded_metadata_20260901_r02.xlsx"
EXPANDED_CSV = OUTPUT_DIR / "mapillary_india_expanded_metadata_20260901_r02.csv"

TARGET_ORDER = ("stop", "maximum_speed_limit_50_km_h", "no_left_turn")
SHEET_NAMES = {
    "stop": "STOP",
    "maximum_speed_limit_50_km_h": "Speed 50",
    "no_left_turn": "No Left Turn",
}
LOCAL_ACCEPTED = {
    "stop": (0, 0),
    "maximum_speed_limit_50_km_h": (0, 0),
    "no_left_turn": (1, 1),
}

ADDITIONAL_CITIES = (
    ("Ghaziabad", 28.6692, 77.4538),
    ("Meerut", 28.9845, 77.7064),
    ("Aligarh", 27.8974, 78.0880),
    ("Bareilly", 28.3670, 79.4304),
    ("Moradabad", 28.8386, 78.7733),
    ("Gorakhpur", 26.7606, 83.3732),
    ("Jhansi", 25.4484, 78.5685),
    ("Gwalior", 26.2183, 78.1828),
    ("Jabalpur", 23.1815, 79.9864),
    ("Ujjain", 23.1765, 75.7885),
    ("Sagar", 23.8388, 78.7378),
    ("Kota", 25.2138, 75.8648),
    ("Ajmer", 26.4499, 74.6399),
    ("Udaipur", 24.5854, 73.7125),
    ("Bikaner", 28.0229, 73.3119),
    ("Rohtak", 28.8955, 76.6066),
    ("Panipat", 29.3909, 76.9635),
    ("Karnal", 29.6857, 76.9905),
    ("Jalandhar", 31.3260, 75.5762),
    ("Patiala", 30.3398, 76.3869),
    ("Shimla", 31.1048, 77.1734),
    ("Haridwar", 29.9457, 78.1642),
    ("Haldwani", 29.2183, 79.5130),
    ("Siliguri", 26.7271, 88.3953),
    ("Durgapur", 23.5204, 87.3119),
    ("Asansol", 23.6739, 86.9524),
    ("Howrah", 22.5958, 88.2636),
    ("Gaya", 24.7914, 85.0002),
    ("Muzaffarpur", 26.1197, 85.3910),
    ("Jamshedpur", 22.8046, 86.2029),
    ("Bokaro", 23.6693, 86.1511),
    ("Rourkela", 22.2604, 84.8536),
    ("Cuttack", 20.4625, 85.8830),
    ("Sambalpur", 21.4669, 83.9812),
    ("Bilaspur", 22.0797, 82.1409),
    ("Durg", 21.1904, 81.2849),
    ("Warangal", 17.9689, 79.5941),
    ("Nizamabad", 18.6725, 78.0941),
    ("Guntur", 16.3067, 80.4365),
    ("Nellore", 14.4426, 79.9865),
    ("Tirupati", 13.6288, 79.4192),
    ("Salem", 11.6643, 78.1460),
    ("Tiruchirappalli", 10.7905, 78.7047),
    ("Erode", 11.3410, 77.7172),
    ("Vellore", 12.9165, 79.1325),
    ("Hubballi", 15.3647, 75.1240),
    ("Belagavi", 15.8497, 74.4977),
    ("Davanagere", 14.4644, 75.9218),
    ("Shivamogga", 13.9299, 75.5681),
    ("Kolhapur", 16.7050, 74.2433),
    ("Solapur", 17.6599, 75.9064),
    ("Nanded", 19.1383, 77.3210),
    ("Jalgaon", 21.0077, 75.5626),
    ("Thane", 19.2183, 72.9781),
    ("Navi Mumbai", 19.0330, 73.0297),
    ("Vasai", 19.3919, 72.8397),
    ("Gandhinagar", 23.2156, 72.6369),
    ("Jamnagar", 22.4707, 70.0577),
    ("Bhavnagar", 21.7645, 72.1519),
    ("Anand", 22.5645, 72.9289),
    ("Panaji", 15.4909, 73.8278),
    ("Margao", 15.2832, 73.9862),
    ("Thrissur", 10.5276, 76.2144),
    ("Kollam", 8.8932, 76.6141),
    ("Kannur", 11.8745, 75.3704),
    ("Puducherry", 11.9416, 79.8083),
)

RING_CITIES = frozenset(
    {
        "Delhi",
        "Mumbai",
        "Bengaluru",
        "Chennai",
        "Hyderabad",
        "Kolkata",
        "Pune",
        "Ahmedabad",
        "Jaipur",
        "Lucknow",
    }
)

TERMS_ROWS = (
    {
        "topic": "Public imagery licence guidance",
        "official_url": "https://help.mapillary.com/hc/en-us/articles/115001770409-CC-BY-SA-license-for-open-data",
        "retrieved_utc_date": "2026-09-01",
        "finding": "Official help states Mapillary images are shared under CC BY-SA and permits distribution/modification with attribution.",
        "project_rule": "Keep image link, contributor and licence attribution with every later downloaded image or crop.",
        "verification_status": "public_guidance_verified",
    },
    {
        "topic": "Attribution format",
        "official_url": "https://help.mapillary.com/hc/en-us/articles/115001770409-CC-BY-SA-license-for-open-data",
        "retrieved_utc_date": "2026-09-01",
        "finding": "Official example identifies title, Mapillary image link, contributor and CC BY-SA.",
        "project_rule": "Use a per-image attribution record; verify the applicable CC BY-SA version before download.",
        "verification_status": "public_guidance_verified",
    },
    {
        "topic": "API/download conditions",
        "official_url": "https://help.mapillary.com/hc/en-us/articles/4407521157138-Downloading-map-data-via-the-Mapillary-web-app",
        "retrieved_utc_date": "2026-09-01",
        "finding": "Official download guidance says imagery and map data use remains subject to Terms of Use, particularly Commercial Terms section 12.",
        "project_rule": "Do not download pixels until the logged-in Terms are manually reviewed for this project.",
        "verification_status": "public_guidance_verified_terms_pending",
    },
    {
        "topic": "Logged-in Terms",
        "official_url": "https://www.mapillary.com/terms",
        "retrieved_utc_date": "2026-09-01",
        "finding": "The Terms page returned 'Not Logged In' to the research session; an API token is not treated as a browser Terms session.",
        "project_rule": "User must open the Terms while logged in and confirm academic image download/use before pixel acquisition.",
        "verification_status": "manual_logged_in_confirmation_required",
    },
    {
        "topic": "Derived crops and report images",
        "official_url": "https://help.mapillary.com/hc/en-us/articles/115001770409-CC-BY-SA-license-for-open-data",
        "retrieved_utc_date": "2026-09-01",
        "finding": "Public CC BY-SA guidance supports modification/distribution with attribution, but the inaccessible logged-in Terms may add conditions.",
        "project_rule": "Treat crops and report publication as blocked until Terms are confirmed; preserve attribution and share-alike obligations.",
        "verification_status": "conditional_not_authorized_yet",
    },
    {
        "topic": "Private repository storage",
        "official_url": "https://www.mapillary.com/terms",
        "retrieved_utc_date": "2026-09-01",
        "finding": "Accessible public guidance did not expressly resolve persistent private-repository storage of downloaded pixels.",
        "project_rule": "Do not store pixels in the repository until the logged-in Terms or Mapillary support expressly confirms the workflow.",
        "verification_status": "unresolved_blocker",
    },
)

PLAN_EXTRA_COLUMNS = (
    "acquisition_order",
    "acquisition_status",
    "pixel_download_authorized",
    "selection_basis",
    "attribution_title",
    "attribution_contributor",
    "attribution_source_url",
    "attribution_licence",
    "attribution_text_template",
    "terms_status",
)
EXPANDED_EXTRA_COLUMNS = (
    "discovery_source",
    "search_box_id",
    "search_type",
    "r07_independence_group_id",
)


def main() -> int:
    """Execute frozen planning and expanded metadata-only discovery."""
    try:
        _ensure_inputs_and_unique_outputs()
        r07_rows = _read_csv(R07_CSV)
        _validate_r07_source(r07_rows)
        plan_rows = build_no_parking_acquisition_plan(r07_rows)
        if (
            len(plan_rows) != 64
            or len({row["independence_group_id"] for row in plan_rows}) != 13
        ):
            raise MapillaryMetadataError(
                "Frozen no-parking pool is not exactly 64 rows / 13 groups"
            )

        token = os.getenv("MAPILLARY_ACCESS_TOKEN", "")
        if not token:
            raise MapillaryMetadataError("MAPILLARY_ACCESS_TOKEN is unavailable")
        client = MapillaryClient(token)
        client.verify_authentication()
        print("MAPILLARY_AUTHENTICATION_SUCCESS=True", flush=True)
        retrieval_timestamp = datetime.now(UTC).replace(microsecond=0).isoformat()
        boxes = _expanded_search_boxes()
        print(f"EXPANDED_METADATA_SCAN_BOXES={len(boxes)}", flush=True)
        features, coverage_rows = _discover_features(client, boxes)
        associations = _feature_image_associations(client, features)
        image_ids = sorted({image_id for _, image_id in associations})
        image_metadata = _fetch_image_metadata(client, image_ids)
        new_rows = _build_candidate_rows(
            features, associations, image_metadata, retrieval_timestamp
        )
        for row in new_rows:
            feature = features[str(row["map_feature_id"])]
            row["discovery_source"] = "expanded_20260901"
            row["search_box_id"] = feature["search_box_id"]
            row["search_type"] = feature["search_type"]
        new_rows = _deduplicate_candidate_images(new_rows)
        for row in r07_rows:
            row["discovery_source"] = "r07"
            row["search_box_id"] = "r07_city_centre"
            row["search_type"] = "r07_city_centre"
        combined_rows, duplicate_removals = merge_expanded_candidates(
            r07_rows, new_rows
        )
        summary_rows = _summary_rows(
            combined_rows,
            r07_rows,
            duplicate_removals,
            coverage_rows,
            features,
            associations,
        )
        dependency_rows = _dependency_rows(combined_rows)
        _write_all_outputs(
            plan_rows,
            combined_rows,
            summary_rows,
            dependency_rows,
            coverage_rows,
            retrieval_timestamp,
            client.request_count,
        )
        result = {
            "authentication_success": True,
            "token_serialized_or_logged": False,
            "image_pixels_downloaded": 0,
            "no_parking_plan_rows": len(plan_rows),
            "no_parking_plan_groups": len(
                {row["independence_group_id"] for row in plan_rows}
            ),
            "expanded_search_boxes": len(boxes),
            "expanded_search_errors": sum(
                row["query_status"] != "success" for row in coverage_rows
            ),
            "summary": summary_rows,
            "artifacts": {
                path.name: {"path": str(path), "sha256": _sha256(path)}
                for path in (
                    PLAN_PDF,
                    PLAN_XLSX,
                    PLAN_CSV,
                    EXPANDED_PDF,
                    EXPANDED_XLSX,
                    EXPANDED_CSV,
                )
            },
            "training_evaluation_calibration_prototypes": False,
        }
        print(json.dumps(result, indent=2))
        return 0
    except (KeyError, MapillaryMetadataError, OSError, TypeError, ValueError) as error:
        print(f"FATAL: {error}", file=sys.stderr)
        return 2


def _ensure_inputs_and_unique_outputs() -> None:
    if not R07_CSV.is_file():
        raise MapillaryMetadataError(f"Missing r07 source: {R07_CSV}")
    for path in _output_paths():
        if path.exists():
            raise MapillaryMetadataError(f"Unique output already exists: {path.name}")


def _output_paths() -> tuple[Path, ...]:
    return PLAN_PDF, PLAN_XLSX, PLAN_CSV, EXPANDED_PDF, EXPANDED_XLSX, EXPANDED_CSV


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _validate_r07_source(rows: Sequence[Mapping[str, Any]]) -> None:
    if len(rows) != 88:
        raise MapillaryMetadataError(f"Expected 88 r07 rows, found {len(rows)}")
    counts = Counter(str(row["project_class"]) for row in rows)
    expected = {
        "stop": 15,
        "maximum_speed_limit_50_km_h": 4,
        "no_left_turn": 5,
        "no_parking": 64,
    }
    if dict(counts) != expected:
        raise MapillaryMetadataError(f"Unexpected r07 class counts: {dict(counts)}")
    no_parking_groups = {
        str(row["independence_group_id"])
        for row in rows
        if row["project_class"] == "no_parking"
    }
    if len(no_parking_groups) != 13:
        raise MapillaryMetadataError("Expected 13 frozen no-parking groups")


def _expanded_search_boxes() -> list[dict[str, Any]]:
    boxes: list[dict[str, Any]] = []
    for index, (city, latitude, longitude) in enumerate(ADDITIONAL_CITIES, 1):
        half_span = 0.0075
        boxes.append(
            {
                "search_box_id": f"additional_city_{index:03d}",
                "area_name": city,
                "search_type": "additional_city_or_town",
                "bbox": (
                    longitude - half_span,
                    latitude - half_span,
                    longitude + half_span,
                    latitude + half_span,
                ),
                "india_evidence": f"Curated interior coordinate centred on {city}, India",
            }
        )
    ring_index = 0
    for city, latitude, longitude in CITY_CENTRES:
        if city not in RING_CITIES:
            continue
        for lat_offset in (-0.015, 0.0, 0.015):
            for lon_offset in (-0.015, 0.0, 0.015):
                if lat_offset == 0.0 and lon_offset == 0.0:
                    continue
                ring_index += 1
                half_span = 0.005
                boxes.append(
                    {
                        "search_box_id": f"major_city_ring_{ring_index:03d}",
                        "area_name": f"{city} outer tile",
                        "search_type": "non_overlapping_major_city_ring",
                        "bbox": (
                            longitude + lon_offset - half_span,
                            latitude + lat_offset - half_span,
                            longitude + lon_offset + half_span,
                            latitude + lat_offset + half_span,
                        ),
                        "india_evidence": (
                            f"Non-overlapping outer urban tile around {city}, India; "
                            "excludes the r07 centre box"
                        ),
                    }
                )
    identifiers = [str(box["search_box_id"]) for box in boxes]
    if len(identifiers) != len(set(identifiers)):
        raise MapillaryMetadataError("Expanded search box IDs are not unique")
    return boxes


def _discover_features(
    client: MapillaryClient, boxes: Sequence[Mapping[str, Any]]
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    features: dict[str, dict[str, Any]] = {}
    coverage: list[dict[str, Any]] = []

    def fetch_box(
        box: Mapping[str, Any],
    ) -> tuple[Mapping[str, Any], list[dict[str, Any]]]:
        bbox = ",".join(f"{float(value):.6f}" for value in box["bbox"])
        payload = client.get_json(
            "/map_features",
            {
                "fields": "id,object_type,object_value,geometry",
                "bbox": bbox,
                "limit": "2000",
            },
        )
        data = payload.get("data", [])
        if not isinstance(data, list):
            raise MapillaryMetadataError("Mapillary feature data is not a list")
        return box, [dict(row) for row in data if isinstance(row, dict)]

    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = {executor.submit(fetch_box, box): box for box in boxes}
        completed = 0
        for future in as_completed(futures):
            completed += 1
            original_box = futures[future]
            try:
                box, rows = future.result()
                status = "success"
            except MapillaryMetadataError:
                box, rows, status = original_box, [], "sanitized_api_error"
            exact_count = 0
            duplicate_features = 0
            for row in rows:
                taxonomy = str(row.get("object_value", ""))
                project_class = map_exact_taxonomy_label(taxonomy)
                if (
                    project_class not in EXPANDED_TARGET_CLASSES
                    or row.get("object_type") != "trafficsign"
                ):
                    continue
                exact_count += 1
                feature_id = str(row.get("id", ""))
                coordinates = _coordinates(row.get("geometry"))
                if not re.fullmatch(r"\d+", feature_id) or coordinates is None:
                    continue
                if feature_id in features:
                    duplicate_features += 1
                    continue
                features[feature_id] = {
                    "map_feature_id": feature_id,
                    "exact_mapillary_taxonomy_label": taxonomy,
                    "project_class": project_class,
                    "longitude": coordinates[0],
                    "latitude": coordinates[1],
                    "city_search_area": str(box["area_name"]),
                    "search_box_id": str(box["search_box_id"]),
                    "search_type": str(box["search_type"]),
                    "geographic_evidence_india": str(box["india_evidence"]),
                }
            coverage.append(
                {
                    "search_box_id": box["search_box_id"],
                    "area_name": box["area_name"],
                    "search_type": box["search_type"],
                    "bbox_west_south_east_north": ",".join(
                        f"{float(value):.6f}" for value in box["bbox"]
                    ),
                    "india_evidence": box["india_evidence"],
                    "query_status": status,
                    "api_feature_rows": len(rows),
                    "possible_2000_row_truncation": (
                        "yes" if len(rows) >= 2000 else "no"
                    ),
                    "exact_target_feature_hits": exact_count,
                    "duplicate_feature_hits": duplicate_features,
                }
            )
            if completed % 25 == 0:
                errors = sum(row["query_status"] != "success" for row in coverage)
                print(
                    f"EXPANDED_SCAN_PROGRESS={completed}/{len(boxes)};"
                    f"TARGET_FEATURES={len(features)};ERRORS={errors}",
                    flush=True,
                )
    coverage.sort(key=lambda row: str(row["search_box_id"]))
    errors = sum(row["query_status"] != "success" for row in coverage)
    if errors > max(5, math.ceil(len(boxes) * 0.10)):
        raise MapillaryMetadataError(
            f"Too many expanded search failures: {errors}/{len(boxes)}"
        )
    return features, coverage


def _summary_rows(
    combined_rows: Sequence[Mapping[str, Any]],
    r07_rows: Sequence[Mapping[str, Any]],
    duplicate_removals: int,
    coverage_rows: Sequence[Mapping[str, Any]],
    features: Mapping[str, Mapping[str, Any]],
    associations: Sequence[tuple[str, str]],
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    successful = sum(row["query_status"] == "success" for row in coverage_rows)
    errors = len(coverage_rows) - successful
    associated_features = {feature_id for feature_id, _ in associations}
    for project_class in TARGET_ORDER:
        rows = [row for row in combined_rows if row["project_class"] == project_class]
        old_rows = [row for row in r07_rows if row["project_class"] == project_class]
        new_rows = [row for row in rows if row["discovery_source"] != "r07"]
        groups = {str(row["independence_group_id"]) for row in rows}
        local_photos, local_groups = LOCAL_ACCEPTED[project_class]
        remaining_photos = max(0, 30 - local_photos)
        remaining_groups = max(0, 15 - local_groups)
        result.append(
            {
                "project_class": project_class,
                "r07_metadata_candidates": len(old_rows),
                "new_unique_metadata_candidates": len(new_rows),
                "combined_metadata_candidates": len(rows),
                "combined_conservative_groups": len(groups),
                "current_local_accepted_photos": local_photos,
                "current_local_accepted_groups": local_groups,
                "remaining_photo_target_before_review": remaining_photos,
                "remaining_group_target_before_review": remaining_groups,
                "ready_for_pixel_acquisition": (
                    "yes"
                    if len(rows) >= remaining_photos and len(groups) >= remaining_groups
                    else "no"
                ),
                "exact_taxonomy_labels_observed": ";".join(
                    sorted({str(row["exact_mapillary_taxonomy_label"]) for row in rows})
                ),
                "deduplicated_against_r07_and_expansion": duplicate_removals,
                "expanded_boxes_successful": successful,
                "expanded_box_errors": errors,
                "expanded_exact_features_without_linked_images": sum(
                    feature_id not in associated_features
                    and feature["project_class"] == project_class
                    for feature_id, feature in features.items()
                ),
                "metadata_candidates_are_approved_photos": "no",
            }
        )
    return result


def _dependency_rows(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[(str(row["project_class"]), str(row["independence_group_id"]))].append(
            row
        )
    return [
        {
            "project_class": project_class,
            "independence_group_id": group_id,
            "candidate_image_count": len(group_rows),
            "map_feature_ids": ";".join(
                sorted({str(row["map_feature_id"]) for row in group_rows})
            ),
            "sequence_ids": ";".join(
                sorted(
                    {
                        str(row["sequence_id"])
                        for row in group_rows
                        if row["sequence_id"]
                    }
                )
            ),
            "contributors": ";".join(
                sorted(
                    {
                        str(row["contributor_name"])
                        for row in group_rows
                        if row["contributor_name"]
                    }
                )
            ),
            "cities_or_areas": ";".join(
                sorted({str(row["city_search_area"]) for row in group_rows})
            ),
        }
        for (project_class, group_id), group_rows in sorted(grouped.items())
    ]


def _write_all_outputs(
    plan_rows: Sequence[Mapping[str, Any]],
    expanded_rows: Sequence[Mapping[str, Any]],
    summary_rows: Sequence[Mapping[str, Any]],
    dependency_rows: Sequence[Mapping[str, Any]],
    coverage_rows: Sequence[Mapping[str, Any]],
    retrieval_timestamp: str,
    request_count: int,
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stage = Path(tempfile.mkdtemp(prefix="mapillary-expanded-stage-", dir=OUTPUT_DIR))
    try:
        staged = {path: stage / path.name for path in _output_paths()}
        _write_csv(
            staged[PLAN_CSV], plan_rows, (*CANDIDATE_COLUMNS, *PLAN_EXTRA_COLUMNS)
        )
        _write_plan_workbook(staged[PLAN_XLSX], plan_rows)
        _write_plan_pdf(staged[PLAN_PDF], plan_rows)
        _write_csv(
            staged[EXPANDED_CSV],
            expanded_rows,
            (*CANDIDATE_COLUMNS, *EXPANDED_EXTRA_COLUMNS),
        )
        _write_expanded_workbook(
            staged[EXPANDED_XLSX],
            expanded_rows,
            summary_rows,
            dependency_rows,
            coverage_rows,
        )
        _write_expanded_pdf(
            staged[EXPANDED_PDF],
            summary_rows,
            dependency_rows,
            coverage_rows,
            retrieval_timestamp,
            request_count,
        )
        _validate_staged(staged, plan_rows, expanded_rows)
        for final_path, staged_path in staged.items():
            os.replace(staged_path, final_path)
    finally:
        shutil.rmtree(stage, ignore_errors=True)


def _write_csv(
    path: Path, rows: Sequence[Mapping[str, Any]], columns: Sequence[str]
) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _write_plan_workbook(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    _add_sheet(workbook, "Acquisition Plan", rows)
    group_rows = _dependency_rows(rows)
    _add_sheet(workbook, "Group Summary", group_rows)
    _add_sheet(workbook, "Terms Snapshot", TERMS_ROWS)
    workbook.save(path)


def _write_expanded_workbook(
    path: Path,
    rows: Sequence[Mapping[str, Any]],
    summary: Sequence[Mapping[str, Any]],
    dependencies: Sequence[Mapping[str, Any]],
    coverage: Sequence[Mapping[str, Any]],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_sheet(workbook, "Summary", summary)
    for project_class in TARGET_ORDER:
        _add_sheet(
            workbook,
            SHEET_NAMES[project_class],
            [row for row in rows if row["project_class"] == project_class],
            fallback_columns=(*CANDIDATE_COLUMNS, *EXPANDED_EXTRA_COLUMNS),
        )
    _add_sheet(workbook, "Dependency Groups", dependencies)
    _add_sheet(workbook, "Search Coverage", coverage)
    _add_sheet(workbook, "Terms Snapshot", TERMS_ROWS)
    workbook.save(path)


def _add_sheet(
    workbook: Workbook,
    title: str,
    rows: Sequence[Mapping[str, Any]],
    fallback_columns: Sequence[str] = (),
) -> None:
    sheet = workbook.create_sheet(title)
    columns = list(rows[0].keys()) if rows else list(fallback_columns)
    for column, name in enumerate(columns, 1):
        cell = sheet.cell(1, column, name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row_number, row in enumerate(rows, 2):
        for column, name in enumerate(columns, 1):
            sheet.cell(row_number, column, row.get(name, ""))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, name in enumerate(columns, 1):
        width = min(55, max(12, len(name) + 2))
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_plan_pdf(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph(
            "Frozen no_parking pixel-acquisition plan — metadata only", styles["Title"]
        ),
        Paragraph(
            "The complete r07 no-parking pool is frozen: 64 candidates across 13 conservative dependency groups. All candidates are retained and ordered round-robin across groups to maximize metadata diversity. Pixel download remains unauthorized pending manual logged-in Terms confirmation.",
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
        _table(
            [
                ["Plan rows", "Groups", "Pixels downloaded", "Terms status"],
                [64, 13, 0, "BLOCKED — manual logged-in confirmation required"],
            ],
            [25 * mm, 20 * mm, 30 * mm, 105 * mm],
        ),
        PageBreak(),
    ]
    for start in range(0, len(rows), 12):
        page_rows = rows[start : start + 12]
        story.append(
            Paragraph(
                f"Acquisition candidates {start + 1}–{start + len(page_rows)}",
                styles["Heading1"],
            )
        )
        table_rows = [
            ["Order", "Image ID", "Feature ID", "Group", "City", "Contributor"]
        ]
        table_rows.extend(
            [
                row["acquisition_order"],
                row["mapillary_image_id"],
                row["map_feature_id"],
                row["independence_group_id"],
                row["city_search_area"],
                row["contributor_name"],
            ]
            for row in page_rows
        )
        story.append(
            _table(
                table_rows,
                [14 * mm, 34 * mm, 34 * mm, 38 * mm, 33 * mm, 36 * mm],
                font_size=7,
            )
        )
        if start + 12 < len(rows):
            story.append(PageBreak())
    story.extend(
        [
            PageBreak(),
            Paragraph("Terms and attribution gate", styles["Heading1"]),
            _mapping_table(TERMS_ROWS, "topic", "project_rule"),
            Spacer(1, 4 * mm),
            Paragraph(
                "No pixels may be downloaded until the user manually reviews the logged-in Mapillary Terms and confirms the proposed academic-use, derived-crop, report-image and private-storage workflow.",
                styles["BodyText"],
            ),
        ]
    )
    SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    ).build(story)


def _write_expanded_pdf(
    path: Path,
    summary: Sequence[Mapping[str, Any]],
    dependencies: Sequence[Mapping[str, Any]],
    coverage: Sequence[Mapping[str, Any]],
    retrieval_timestamp: str,
    request_count: int,
) -> None:
    styles = getSampleStyleSheet()
    successful = sum(row["query_status"] == "success" for row in coverage)
    story: list[Any] = [
        Paragraph("Expanded India Mapillary metadata discovery", styles["Title"]),
        Paragraph(
            f"Authenticated metadata-only retrieval completed {retrieval_timestamp}. Search boxes: {len(coverage)}; successful: {successful}; API metadata requests: {request_count}; image pixels downloaded: 0.",
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
        _rows_table(
            summary,
            (
                "project_class",
                "r07_metadata_candidates",
                "new_unique_metadata_candidates",
                "combined_metadata_candidates",
                "combined_conservative_groups",
                "ready_for_pixel_acquisition",
            ),
        ),
        PageBreak(),
        Paragraph("Systematic search coverage", styles["Heading1"]),
        Paragraph(
            "Coverage adds curated Indian cities/towns plus eight non-overlapping outer tiles around each of ten major cities. It excludes every previous r07 centre box and never queries no_parking as an active target. Exact target taxonomies are filtered client-side from traffic-sign map features.",
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
        _rows_table(
            [
                {
                    "search_type": key,
                    "boxes": len(value),
                    "successful": sum(
                        row["query_status"] == "success" for row in value
                    ),
                    "exact_hits": sum(
                        int(row["exact_target_feature_hits"]) for row in value
                    ),
                    "possible_truncations": sum(
                        row["possible_2000_row_truncation"] == "yes" for row in value
                    ),
                }
                for key, value in _group_by(coverage, "search_type").items()
            ],
            (
                "search_type",
                "boxes",
                "successful",
                "exact_hits",
                "possible_truncations",
            ),
        ),
        PageBreak(),
        Paragraph("Exact taxonomy and readiness", styles["Heading1"]),
        _rows_table(
            summary,
            (
                "project_class",
                "exact_taxonomy_labels_observed",
                "remaining_photo_target_before_review",
                "remaining_group_target_before_review",
                "ready_for_pixel_acquisition",
            ),
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "Readiness is metadata availability only. No candidate is visually approved, and no metadata count is added to the approved-photograph totals.",
            styles["BodyText"],
        ),
        PageBreak(),
        Paragraph("Conservative dependency groups", styles["Heading1"]),
        _rows_table(
            [
                {
                    "project_class": key,
                    "groups": len(value),
                    "candidate_images": sum(
                        int(row["candidate_image_count"]) for row in value
                    ),
                    "multi_frame_groups": sum(
                        int(row["candidate_image_count"]) > 1 for row in value
                    ),
                }
                for key, value in _group_by(dependencies, "project_class").items()
            ],
            ("project_class", "groups", "candidate_images", "multi_frame_groups"),
        ),
        Paragraph(
            "Grouping joins shared map-feature IDs, shared image evidence, signs within 20 metres, and signs within 75 metres when contributor/sequence and capture-time evidence also agree. Repeated frames never count as independent.",
            styles["BodyText"],
        ),
        PageBreak(),
        Paragraph("Terms, attribution and storage gate", styles["Heading1"]),
        _mapping_table(TERMS_ROWS, "topic", "verification_status"),
        Spacer(1, 4 * mm),
        Paragraph(
            "The public CC BY-SA and attribution guidance is recorded. The logged-in Terms page was not accessible to this research session, so pixel download, derived crops, report-image publication and persistent private-repository storage remain blocked pending manual verification.",
            styles["BodyText"],
        ),
        PageBreak(),
        Paragraph("Exact next action", styles["Heading1"]),
        Paragraph(
            "Review the combined candidate and dependency-group tables. For any class marked ready, manually review the logged-in Terms and confirm the attribution/storage workflow. Only then may a separate explicitly approved phase request image pixels. Classes still below 30 metadata candidates or 15 conservative groups require another acquisition source or user-captured photographs; quality thresholds must not be lowered.",
            styles["BodyText"],
        ),
    ]
    SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    ).build(story)


def _validate_staged(
    staged: Mapping[Path, Path],
    plan_rows: Sequence[Mapping[str, Any]],
    expanded_rows: Sequence[Mapping[str, Any]],
) -> None:
    for path in staged.values():
        if not path.is_file() or path.stat().st_size == 0:
            raise MapillaryMetadataError(
                f"Missing or empty staged artifact: {path.name}"
            )
    plan_csv = _read_csv(staged[PLAN_CSV])
    expanded_csv = _read_csv(staged[EXPANDED_CSV])
    if [row["mapillary_image_id"] for row in plan_csv] != [
        str(row["mapillary_image_id"]) for row in plan_rows
    ]:
        raise MapillaryMetadataError("No-parking plan CSV IDs differ from source rows")
    if [row["mapillary_image_id"] for row in expanded_csv] != [
        str(row["mapillary_image_id"]) for row in expanded_rows
    ]:
        raise MapillaryMetadataError("Expanded CSV IDs differ from combined rows")
    plan_workbook = load_workbook(staged[PLAN_XLSX], read_only=True, data_only=True)
    try:
        if plan_workbook.sheetnames != [
            "Acquisition Plan",
            "Group Summary",
            "Terms Snapshot",
        ]:
            raise MapillaryMetadataError("Unexpected no-parking workbook sheets")
    finally:
        plan_workbook.close()
    expanded_workbook = load_workbook(
        staged[EXPANDED_XLSX], read_only=True, data_only=True
    )
    try:
        if expanded_workbook.sheetnames != [
            "Summary",
            "STOP",
            "Speed 50",
            "No Left Turn",
            "Dependency Groups",
            "Search Coverage",
            "Terms Snapshot",
        ]:
            raise MapillaryMetadataError("Unexpected expanded workbook sheets")
    finally:
        expanded_workbook.close()
    if len(PdfReader(str(staged[PLAN_PDF])).pages) < 2:
        raise MapillaryMetadataError("No-parking PDF is unexpectedly short")
    if len(PdfReader(str(staged[EXPANDED_PDF])).pages) < 5:
        raise MapillaryMetadataError("Expanded PDF is unexpectedly short")
    if any(row["pixel_download_authorized"] != "no" for row in plan_rows):
        raise MapillaryMetadataError("Plan incorrectly authorizes pixel download")
    if any(row["project_class"] == "no_parking" for row in expanded_rows):
        raise MapillaryMetadataError(
            "Expanded output contains prohibited no_parking rows"
        )


def _rows_table(rows: Sequence[Mapping[str, Any]], columns: Sequence[str]) -> Table:
    data = [[_header(column) for column in columns]]
    data.extend([[str(row.get(column, "")) for column in columns] for row in rows])
    width = 265 * mm / max(1, len(columns))
    return _table(data, [width] * len(columns), font_size=7)


def _mapping_table(rows: Sequence[Mapping[str, Any]], left: str, right: str) -> Table:
    data = [[_header(left), _header(right)]]
    data.extend([[str(row[left]), str(row[right])] for row in rows])
    return _table(data, [70 * mm, 190 * mm], font_size=7)


def _table(
    data: Sequence[Sequence[Any]], widths: Sequence[float], font_size: int = 8
) -> Table:
    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#EAF2F8")],
                ),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _group_by(
    rows: Sequence[Mapping[str, Any]], key: str
) -> dict[str, list[Mapping[str, Any]]]:
    result: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        result[str(row[key])].append(row)
    return dict(sorted(result.items()))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _header(value: str) -> str:
    return value.replace("_", " ").title()


if __name__ == "__main__":
    raise SystemExit(main())
