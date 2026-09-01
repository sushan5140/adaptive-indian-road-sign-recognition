"""Finalize V2 review, viability, group-safe splits, and review reports."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import sys
import tempfile
from collections import Counter
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from data.v2_finalization import (  # noqa: E402
    SPLIT_MANIFEST_COLUMNS,
    SPLIT_ORDER,
    V2FinalizationError,
    analyze_class_viability,
    apply_final_review_decisions,
    group_safe_split,
    validate_approved_set,
)
from data.v2_review import V2_REVIEW_COLUMNS  # noqa: E402

FINAL_STATUS_COUNTS = {
    "approved": 412,
    "rejected": 118,
    "relabel": 0,
    "pending": 0,
}
REQUIRED_WORKBOOK_SHEETS = (
    "Review Summary",
    "Class Viability",
    "Train Split",
    "Validation Split",
    "Test Split",
    "Leakage Checks",
)


def build_parser() -> argparse.ArgumentParser:
    """Build command-line arguments for V2 finalization."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source-review", default="outputs/v2_review/dataset_b_v2_review.csv"
    )
    parser.add_argument(
        "--decision-csv",
        required=True,
    )
    parser.add_argument(
        "--decision-xlsx",
        required=True,
    )
    parser.add_argument(
        "--image-inventory", default="outputs/dataset_b_audit/image_inventory.csv"
    )
    parser.add_argument(
        "--near-duplicate-pairs",
        default="outputs/dataset_b_audit/near_duplicate_pairs.csv",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    """Run the fail-closed V2 final review and split workflow without training."""
    arguments = build_parser().parse_args(argv)
    try:
        source_path = _project_path(arguments.source_review)
        decision_csv = _project_path(arguments.decision_csv)
        decision_xlsx = _project_path(arguments.decision_xlsx)
        source_rows, source_columns = _read_csv(source_path)
        decision_rows, decision_columns = _read_csv(decision_csv)
        xlsx_rows, xlsx_columns = _read_xlsx(decision_xlsx)
        if source_columns != tuple(V2_REVIEW_COLUMNS):
            raise V2FinalizationError("Source review columns differ from V2 schema")
        if decision_columns != source_columns or xlsx_columns != source_columns:
            raise V2FinalizationError(
                "Decision source columns differ from review source"
            )
        if decision_rows != xlsx_rows:
            raise V2FinalizationError("Completed CSV and XLSX decision rows differ")
        pre_counts = _status_counts(source_rows)
        if pre_counts != {
            "approved": 318,
            "rejected": 99,
            "relabel": 0,
            "pending": 113,
        }:
            raise V2FinalizationError(
                f"Unexpected pre-finalization status counts: {pre_counts}"
            )
        result = apply_final_review_decisions(source_rows, decision_rows)
        if dict(result.status_counts) != FINAL_STATUS_COUNTS or len(result.rows) != 530:
            raise V2FinalizationError(
                f"Unexpected final review state: rows={len(result.rows)}, "
                f"counts={dict(result.status_counts)}"
            )
        viability = analyze_class_viability(result.rows)
        viable_classes = tuple(
            row["proposed_class"] for row in viability if row["viable"]
        )
        excluded_classes = tuple(
            {
                "proposed_class": row["proposed_class"],
                "reason": row["exclusion_reason"],
            }
            for row in viability
            if not row["viable"]
        )
        inventory_rows, _ = _read_csv(_project_path(arguments.image_inventory))
        image_hashes = {row["image_name"]: row["sha256"] for row in inventory_rows}
        pair_rows, _ = _read_csv(_project_path(arguments.near_duplicate_pairs))
        pairs = tuple((row["left_image"], row["right_image"]) for row in pair_rows)
        approved_checks = validate_approved_set(
            result.rows, image_hashes=image_hashes, near_duplicate_pairs=pairs
        )
        if not approved_checks["all_checks_passed"]:
            raise V2FinalizationError(
                f"Approved-set leakage validation failed: {approved_checks}"
            )
        split_result = group_safe_split(result.rows, viable_classes, random_seed=42)
        class_mapping = {
            class_name: index for index, class_name in enumerate(sorted(viable_classes))
        }
        final_summary = {
            "review_complete": True,
            "total_candidate_rows": len(result.rows),
            "status_counts": dict(result.status_counts),
            "decision_source": {
                "selected": decision_csv.name,
                "csv_sha256": _sha256(decision_csv),
                "xlsx_sha256": _sha256(decision_xlsx),
                "csv_xlsx_rows_identical": True,
                "decision_rows": len(decision_rows),
                "protected_field_differences": [],
                "invalid_or_mismatched_review_ids": [],
                "previously_completed_decisions_preserved": 417,
            },
            "viability_rule": {
                "minimum_approved_photographs": 10,
                "minimum_independent_perceptual_groups": 8,
            },
            "class_viability": viability,
            "viable_classes": list(viable_classes),
            "excluded_classes": list(excluded_classes),
            "approved_set_checks": approved_checks,
        }
        split_summary = {
            **dict(split_result.summary),
            "retained_class_count": len(viable_classes),
            "retained_classes": list(sorted(viable_classes)),
            "excluded_classes": list(excluded_classes),
            "class_mapping": class_mapping,
            "manifest_columns": list(SPLIT_MANIFEST_COLUMNS),
        }
        destinations = _destinations(source_path)
        existing = [
            str(path)
            for key, path in destinations.items()
            if key != "review" and path.exists()
        ]
        if existing:
            raise V2FinalizationError(
                f"Refusing to overwrite final artifacts: {existing}"
            )
        work_root = PROJECT_ROOT / "work"
        work_root.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(
            prefix="v2-finalize-", dir=work_root
        ) as temporary:
            stage = Path(temporary)
            staged = {key: stage / path.name for key, path in destinations.items()}
            _write_csv(staged["review"], result.rows, tuple(V2_REVIEW_COLUMNS))
            _write_json(staged["review_summary"], final_summary)
            _write_json(staged["class_mapping"], class_mapping)
            for split_name in SPLIT_ORDER:
                _write_csv(
                    staged[split_name],
                    split_result.splits[split_name],
                    SPLIT_MANIFEST_COLUMNS,
                )
            _write_json(staged["split_summary"], split_summary)
            pdf_pages = _write_pdf_report(staged["pdf"], final_summary, split_summary)
            _write_xlsx_report(
                staged["xlsx"], final_summary, split_summary, split_result.splits
            )
            _validate_staged(
                staged,
                result.rows,
                split_result.splits,
                class_mapping,
                pdf_pages,
            )
            for destination in destinations.values():
                destination.parent.mkdir(parents=True, exist_ok=True)
            for key, destination in destinations.items():
                os.replace(staged[key], destination)
        output = {
            "status_counts": dict(result.status_counts),
            "viable_classes": list(sorted(viable_classes)),
            "excluded_classes": list(excluded_classes),
            "split_sizes": split_summary["sizes"],
            "per_class_split_counts": split_summary["per_class_sample_counts"],
            "leakage_checks": split_summary["leakage_checks"],
            "pdf_page_count": pdf_pages,
            "xlsx_sheet_names": list(REQUIRED_WORKBOOK_SHEETS),
            "artifacts": {
                key: {"path": str(path), "sha256": _sha256(path)}
                for key, path in destinations.items()
            },
            "training_started": False,
        }
        print(json.dumps(output, ensure_ascii=False, indent=2))
        return 0
    except (OSError, csv.Error, ValueError, V2FinalizationError) as error:
        print(f"FATAL: {error}", file=sys.stderr)
        return 2


def _destinations(source_review: Path) -> dict[str, Path]:
    return {
        "review": source_review,
        "review_summary": PROJECT_ROOT / "outputs/v2_review/final_review_summary.json",
        "class_mapping": PROJECT_ROOT / "outputs/manifests/v2_class_mapping.json",
        "train": PROJECT_ROOT / "outputs/manifests/v2_train.csv",
        "validation": PROJECT_ROOT / "outputs/manifests/v2_validation.csv",
        "test": PROJECT_ROOT / "outputs/manifests/v2_test.csv",
        "split_summary": PROJECT_ROOT / "outputs/manifests/v2_split_summary.json",
        "pdf": PROJECT_ROOT / "outputs/v2_review/v2_final_review_report.pdf",
        "xlsx": PROJECT_ROOT / "outputs/v2_review/v2_final_review_summary.xlsx",
    }


def _read_csv(path: Path) -> tuple[list[dict[str, str]], tuple[str, ...]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise V2FinalizationError(f"CSV has no header: {path}")
        return [dict(row) for row in reader], tuple(reader.fieldnames)


def _read_xlsx(path: Path) -> tuple[list[dict[str, str]], tuple[str, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    values = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not values:
        raise V2FinalizationError(f"XLSX is empty: {path}")
    columns = tuple("" if value is None else str(value) for value in values[0])
    rows = [
        {
            column: "" if value is None else str(value)
            for column, value in zip(columns, row, strict=True)
        }
        for row in values[1:]
    ]
    return rows, columns


def _write_csv(
    path: Path, rows: Iterable[Mapping[str, Any]], columns: tuple[str, ...]
) -> None:
    with path.open("x", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def _write_xlsx_report(
    path: Path,
    final_summary: Mapping[str, Any],
    split_summary: Mapping[str, Any],
    splits: Mapping[str, tuple[dict[str, str], ...]],
) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    review = workbook.create_sheet("Review Summary")
    review.append(("Metric", "Value"))
    for status, count in final_summary["status_counts"].items():
        review.append((status, count))
    review.append(("total_candidate_rows", final_summary["total_candidate_rows"]))
    review.append(("viable_class_count", len(final_summary["viable_classes"])))
    review.append(("excluded_class_count", len(final_summary["excluded_classes"])))
    review.append(("random_seed", split_summary["random_seed"]))
    review.append(("target_ratios", json.dumps(split_summary["target_ratios"])))
    review.append(("actual_ratios", json.dumps(split_summary["actual_ratios"])))
    viability_sheet = workbook.create_sheet("Class Viability")
    viability_columns = (
        "proposed_class",
        "candidate_count",
        "approved_count",
        "rejected_count",
        "approved_perceptual_group_count",
        "viable",
        "exclusion_reason",
    )
    viability_sheet.append(viability_columns)
    for row in final_summary["class_viability"]:
        viability_sheet.append(tuple(row[column] for column in viability_columns))
    split_sheet_names = {
        "train": "Train Split",
        "validation": "Validation Split",
        "test": "Test Split",
    }
    for split_name in SPLIT_ORDER:
        sheet = workbook.create_sheet(split_sheet_names[split_name])
        sheet.append(SPLIT_MANIFEST_COLUMNS)
        for row in splits[split_name]:
            sheet.append(tuple(row[column] for column in SPLIT_MANIFEST_COLUMNS))
    leakage = workbook.create_sheet("Leakage Checks")
    leakage.append(("Check", "Passed", "Details"))
    approved_checks = final_summary["approved_set_checks"]
    for key in (
        "duplicate_review_ids",
        "duplicate_source_image_ids",
        "missing_image_hash_source_ids",
        "exact_duplicate_groups",
        "near_duplicate_pair_group_inconsistencies",
        "cross_label_perceptual_group_conflicts",
    ):
        leakage.append(
            (key, not approved_checks[key], json.dumps(approved_checks[key]))
        )
    for key in (
        "perceptual_group_overlap",
        "source_image_overlap",
        "review_id_overlap",
    ):
        value = split_summary["leakage_checks"][key]
        leakage.append((f"split_{key}", not value, json.dumps(value)))
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            length = min(
                60,
                max(
                    12,
                    max(len(str(cell.value or "")) for cell in column_cells) + 2,
                ),
            )
            sheet.column_dimensions[column_cells[0].column_letter].width = length
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def _write_pdf_report(
    path: Path, final_summary: Mapping[str, Any], split_summary: Mapping[str, Any]
) -> int:
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], alignment=TA_CENTER, fontSize=20
    )
    heading = styles["Heading2"]
    body = styles["BodyText"]
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Baseline V2 Final Review Report",
        author="Adaptive Indian Road Sign Recognition project",
    )
    story: list[Any] = [
        Paragraph("Baseline V2 Final Review Report", title),
        Spacer(1, 5 * mm),
        Paragraph("Final review counts", heading),
        _styled_table(
            [["Status", "Count"]]
            + [
                [status, count]
                for status, count in final_summary["status_counts"].items()
            ],
            [55 * mm, 30 * mm],
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "Viability rule: at least 10 approved photographs and at least 8 "
            "independent approved perceptual groups. The rule was not changed.",
            body,
        ),
        Spacer(1, 4 * mm),
        Paragraph("Per-class viability", heading),
    ]
    viability_data = [
        ["Class", "Candidates", "Approved", "Rejected", "Groups", "Viable", "Reason"]
    ]
    for row in final_summary["class_viability"]:
        viability_data.append(
            [
                row["proposed_class"],
                row["candidate_count"],
                row["approved_count"],
                row["rejected_count"],
                row["approved_perceptual_group_count"],
                "yes" if row["viable"] else "no",
                row["exclusion_reason"] or "—",
            ]
        )
    story.extend(
        [
            _styled_table(
                viability_data,
                [55 * mm, 20 * mm, 20 * mm, 20 * mm, 20 * mm, 17 * mm, 95 * mm],
                font_size=7.5,
            ),
            PageBreak(),
            Paragraph("Group-safe split", heading),
            Paragraph(split_summary["algorithm"], body),
            Spacer(1, 3 * mm),
            _styled_table(
                [["Split", "Samples", "Target ratio", "Actual ratio", "Deviation"]]
                + [
                    [
                        split_name,
                        split_summary["sizes"][split_name],
                        f"{split_summary['target_ratios'][split_name]:.2%}",
                        f"{split_summary['actual_ratios'][split_name]:.2%}",
                        f"{split_summary['ratio_deviations'][split_name]:+.2%}",
                    ]
                    for split_name in SPLIT_ORDER
                ],
                [40 * mm, 30 * mm, 35 * mm, 35 * mm, 35 * mm],
            ),
            Spacer(1, 4 * mm),
            Paragraph("Per-class split counts", heading),
        ]
    )
    per_class_data = [["Class", "Train", "Validation", "Test", "Total"]]
    for class_name, counts in split_summary["per_class_sample_counts"].items():
        per_class_data.append(
            [
                class_name,
                counts["train"],
                counts["validation"],
                counts["test"],
                sum(counts.values()),
            ]
        )
    story.extend(
        [
            _styled_table(
                per_class_data,
                [80 * mm, 28 * mm, 28 * mm, 28 * mm, 28 * mm],
                font_size=8,
            ),
            PageBreak(),
            Paragraph("Leakage and safety checks", heading),
        ]
    )
    approved_checks = final_summary["approved_set_checks"]
    leakage_rows = [["Check", "Result", "Details"]]
    for key in (
        "duplicate_review_ids",
        "duplicate_source_image_ids",
        "exact_duplicate_groups",
        "near_duplicate_pair_group_inconsistencies",
        "cross_label_perceptual_group_conflicts",
    ):
        leakage_rows.append(
            [
                key,
                "PASS" if not approved_checks[key] else "FAIL",
                json.dumps(approved_checks[key]),
            ]
        )
    for key in (
        "perceptual_group_overlap",
        "source_image_overlap",
        "review_id_overlap",
    ):
        value = split_summary["leakage_checks"][key]
        leakage_rows.append(
            [f"split_{key}", "PASS" if not value else "FAIL", json.dumps(value)]
        )
    story.extend(
        [
            _styled_table(leakage_rows, [90 * mm, 25 * mm, 130 * mm], font_size=8),
            Spacer(1, 5 * mm),
            Paragraph("Excluded classes", heading),
            _styled_table(
                [["Class", "Reason"]]
                + [
                    [row["proposed_class"], row["reason"]]
                    for row in final_summary["excluded_classes"]
                ],
                [75 * mm, 170 * mm],
            ),
            Spacer(1, 5 * mm),
            Paragraph(
                "Test-use policy: the test split is held out and must not be used "
                "for model selection. Validation macro F1 selects checkpoints. No "
                "model training or evaluation was performed in this phase.",
                body,
            ),
        ]
    )
    document.build(story, onFirstPage=_page_number, onLaterPages=_page_number)
    return len(PdfReader(str(path)).pages)


def _styled_table(
    data: list[list[Any]], widths: list[float], *, font_size: float = 8.5
) -> Table:
    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), font_size),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#EAF2F8")],
                ),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _page_number(canvas: Any, document: Any) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(
        landscape(A4)[0] - 12 * mm,
        7 * mm,
        f"Page {document.page}",
    )
    canvas.restoreState()


def _validate_staged(
    staged: Mapping[str, Path],
    final_rows: tuple[dict[str, str], ...],
    splits: Mapping[str, tuple[dict[str, str], ...]],
    class_mapping: Mapping[str, int],
    pdf_pages: int,
) -> None:
    review_rows, review_columns = _read_csv(staged["review"])
    if review_columns != tuple(V2_REVIEW_COLUMNS) or review_rows != list(final_rows):
        raise V2FinalizationError("Staged final review differs from validated rows")
    if _status_counts(review_rows) != FINAL_STATUS_COUNTS:
        raise V2FinalizationError("Staged final review counts are incorrect")
    manifest_ids: set[str] = set()
    group_owners: dict[str, set[str]] = {}
    source_owners: dict[str, set[str]] = {}
    for split_name in SPLIT_ORDER:
        rows, columns = _read_csv(staged[split_name])
        if columns != SPLIT_MANIFEST_COLUMNS or rows != list(splits[split_name]):
            raise V2FinalizationError(f"Staged {split_name} manifest differs")
        if {row["class_name"] for row in rows} != set(class_mapping):
            raise V2FinalizationError(f"Staged {split_name} lacks retained classes")
        for row in rows:
            if row["review_id"] in manifest_ids:
                raise V2FinalizationError("Review ID appears in multiple manifests")
            manifest_ids.add(row["review_id"])
            group_owners.setdefault(row["perceptual_group_id"], set()).add(split_name)
            source_owners.setdefault(row["source_image_id"], set()).add(split_name)
    if len(manifest_ids) != 412:
        raise V2FinalizationError(f"Expected 412 split rows, found {len(manifest_ids)}")
    if any(len(owners) > 1 for owners in group_owners.values()):
        raise V2FinalizationError("A perceptual group crosses staged manifests")
    if any(len(owners) > 1 for owners in source_owners.values()):
        raise V2FinalizationError("A source image crosses staged manifests")
    workbook = load_workbook(staged["xlsx"], read_only=True, data_only=False)
    if tuple(workbook.sheetnames) != REQUIRED_WORKBOOK_SHEETS:
        workbook.close()
        raise V2FinalizationError("Final XLSX sheet names differ from requirements")
    workbook.close()
    reader = PdfReader(str(staged["pdf"]))
    if len(reader.pages) != pdf_pages or pdf_pages < 1:
        raise V2FinalizationError("Final PDF page count is invalid")
    pdf_text = "\n".join(page.extract_text() or "" for page in reader.pages)
    for required_text in (
        "Baseline V2 Final Review Report",
        "Per-class viability",
        "Group-safe split",
        "Leakage and safety checks",
        "Excluded classes",
    ):
        if required_text not in pdf_text:
            raise V2FinalizationError(f"Final PDF is missing section {required_text!r}")


def _status_counts(rows: Iterable[Mapping[str, str]]) -> dict[str, int]:
    counts = Counter(row["review_status"].strip().casefold() for row in rows)
    return {
        status: counts[status]
        for status in ("approved", "rejected", "relabel", "pending")
    }


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _project_path(value: str) -> Path:
    path = Path(value).expanduser()
    return path.resolve() if path.is_absolute() else (PROJECT_ROOT / path).resolve()


if __name__ == "__main__":
    raise SystemExit(main())
