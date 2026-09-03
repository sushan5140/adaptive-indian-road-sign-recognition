"""Discover India-only Mapillary traffic-sign metadata without image pixels."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
import sys
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping, Sequence
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from data.mapillary_metadata import (  # noqa: E402
    EXACT_TAXONOMY_PREFIXES,
    MapillaryMetadataError,
    assign_conservative_independence_groups,
    map_exact_taxonomy_label,
)

OUTPUT_DIR = PROJECT_ROOT / "outputs/open_set_acquisition"
PDF_PATH = OUTPUT_DIR / "mapillary_india_metadata_report_20260901_r07.pdf"
XLSX_PATH = OUTPUT_DIR / "mapillary_india_metadata_20260901_r07.xlsx"
CSV_PATH = OUTPUT_DIR / "mapillary_india_metadata_20260901_r07.csv"

API_ROOT = "https://graph.mapillary.com"
USER_AGENT = "adaptive-road-sign-research-metadata/1.0"
TARGET_ORDER = (
    "stop",
    "maximum_speed_limit_50_km_h",
    "no_left_turn",
    "no_parking",
)
SHEET_NAMES = {
    "stop": "STOP",
    "maximum_speed_limit_50_km_h": "Speed 50",
    "no_left_turn": "No Left Turn",
    "no_parking": "No Parking",
}
LOCAL_STATE = {
    "stop": (0, 0),
    "maximum_speed_limit_50_km_h": (0, 0),
    "no_left_turn": (1, 1),
    "no_parking": (4, 3),
}
CITY_CENTRES = (
    ("Delhi", 28.6139, 77.2090),
    ("Mumbai", 19.0760, 72.8777),
    ("Bengaluru", 12.9716, 77.5946),
    ("Chennai", 13.0827, 80.2707),
    ("Hyderabad", 17.3850, 78.4867),
    ("Kolkata", 22.5726, 88.3639),
    ("Pune", 18.5204, 73.8567),
    ("Ahmedabad", 23.0225, 72.5714),
    ("Jaipur", 26.9124, 75.7873),
    ("Lucknow", 26.8467, 80.9462),
    ("Kanpur", 26.4499, 80.3319),
    ("Nagpur", 21.1458, 79.0882),
    ("Indore", 22.7196, 75.8577),
    ("Bhopal", 23.2599, 77.4126),
    ("Surat", 21.1702, 72.8311),
    ("Vadodara", 22.3072, 73.1812),
    ("Nashik", 19.9975, 73.7898),
    ("Patna", 25.5941, 85.1376),
    ("Ranchi", 23.3441, 85.3096),
    ("Bhubaneswar", 20.2961, 85.8245),
    ("Guwahati", 26.1445, 91.7362),
    ("Chandigarh", 30.7333, 76.7794),
    ("Amritsar", 31.6340, 74.8723),
    ("Ludhiana", 30.9010, 75.8573),
    ("Dehradun", 30.3165, 78.0322),
    ("Jammu", 32.7266, 74.8570),
    ("Kochi", 9.9312, 76.2673),
    ("Thiruvananthapuram", 8.5241, 76.9366),
    ("Kozhikode", 11.2588, 75.7804),
    ("Coimbatore", 11.0168, 76.9558),
    ("Madurai", 9.9252, 78.1198),
    ("Visakhapatnam", 17.6868, 83.2185),
    ("Vijayawada", 16.5062, 80.6480),
    ("Mysuru", 12.2958, 76.6394),
    ("Mangaluru", 12.9141, 74.8560),
    ("Raipur", 21.2514, 81.6296),
    ("Jodhpur", 26.2389, 73.0243),
    ("Agra", 27.1767, 78.0081),
    ("Varanasi", 25.3176, 82.9739),
    ("Prayagraj", 25.4358, 81.8463),
    ("Rajkot", 22.3039, 70.8022),
    ("Aurangabad", 19.8762, 75.3433),
    ("Gurugram", 28.4595, 77.0266),
    ("Noida", 28.5355, 77.3910),
    ("Faridabad", 28.4089, 77.3178),
)
CANDIDATE_COLUMNS = (
    "mapillary_image_id",
    "map_feature_id",
    "all_linked_map_feature_ids",
    "exact_mapillary_taxonomy_label",
    "project_class",
    "sequence_id",
    "capture_timestamp_utc",
    "captured_at_epoch_ms",
    "contributor_id",
    "contributor_name",
    "latitude",
    "longitude",
    "geographic_evidence_india",
    "city_search_area",
    "source_reference",
    "api_retrieval_timestamp_utc",
    "taxonomy_mapping_confidence",
    "independence_group_id",
    "review_status",
    "approved_photograph",
)

TERMS_ROWS = (
    {
        "topic": "Imagery licence",
        "finding": "Official Mapillary help states all Mapillary images are shared under CC BY-SA.",
        "reference": "https://help.mapillary.com/hc/en-us/articles/115001770409-CC-BY-SA-license-for-open-data",
        "retrieval_date": "2026-09-01",
        "project_implication": "Later image use requires attribution and share-alike compliance; no pixels were downloaded in this phase.",
    },
    {
        "topic": "Attribution",
        "finding": "Official example uses image title/link, contributor username/profile link, and CC BY-SA notice.",
        "reference": "https://help.mapillary.com/hc/en-us/articles/115001770409-CC-BY-SA-license-for-open-data",
        "retrieval_date": "2026-09-01",
        "project_implication": "Retain image ID, source link, and contributor metadata before any pixel acquisition.",
    },
    {
        "topic": "Map data access",
        "finding": "Official guidance supports traffic-sign map features through the API and requires compliance with Terms and Commercial Terms.",
        "reference": "https://help.mapillary.com/hc/en-us/articles/4407521157138-Downloading-map-data-via-the-Mapillary-web-app",
        "retrieval_date": "2026-09-01",
        "project_implication": "Review logged-in current terms before image download or repository/report redistribution decisions.",
    },
    {
        "topic": "Repository and report storage",
        "finding": "No blanket conclusion was recorded for repository storage or report examples beyond CC BY-SA guidance.",
        "reference": "https://www.mapillary.com/terms",
        "retrieval_date": "2026-09-01",
        "project_implication": "Treat download/storage/report use as pending a deliberate terms review and attribution plan.",
    },
)


class MapillaryClient:
    """Header-authenticated metadata-only Mapillary Graph API client."""

    def __init__(self, access_token: str) -> None:
        if not access_token:
            raise MapillaryMetadataError("Mapillary access token is unavailable")
        self._headers = {
            "Authorization": f"OAuth {access_token}",
            "User-Agent": USER_AGENT,
        }
        self._lock = threading.Lock()
        self.request_count = 0

    def get_json(
        self, path: str, params: Mapping[str, str] | None = None
    ) -> dict[str, Any]:
        """Retrieve JSON metadata with bounded retry and sanitized errors."""
        if not path.startswith("/") or "access_token" in path:
            raise MapillaryMetadataError("Unsafe Mapillary API path")
        query = urllib.parse.urlencode(params or {})
        url = f"{API_ROOT}{path}" + (f"?{query}" if query else "")
        for attempt in range(1):
            request = urllib.request.Request(url, headers=self._headers)
            try:
                with urllib.request.urlopen(request, timeout=45) as response:
                    payload = json.load(response)
                with self._lock:
                    self.request_count += 1
                if not isinstance(payload, dict):
                    raise MapillaryMetadataError("Mapillary returned non-object JSON")
                return payload
            except urllib.error.HTTPError as error:
                if error.code not in {429, 500, 502, 503, 504} or attempt == 0:
                    raise MapillaryMetadataError(
                        f"Mapillary metadata request failed with HTTP {error.code}"
                    ) from None
            except (TimeoutError, urllib.error.URLError):
                if attempt == 0:
                    raise MapillaryMetadataError(
                        "Mapillary metadata request failed due to network timeout"
                    ) from None
            time.sleep(1.5 * (2**attempt))
        raise MapillaryMetadataError("Mapillary metadata retry loop exhausted")

    def verify_authentication(self) -> None:
        """Verify authentication without requesting any image content fields."""
        payload = self.get_json(
            "/images",
            {
                "fields": "id",
                "bbox": "77.2000,28.6000,77.2001,28.6001",
                "limit": "1",
            },
        )
        if "data" not in payload:
            raise MapillaryMetadataError("Authentication response lacked data field")


def main() -> int:
    """Run authenticated India-only metadata discovery and package results."""
    token = os.getenv("MAPILLARY_ACCESS_TOKEN", "")
    try:
        _ensure_unique_outputs()
        client = MapillaryClient(token)
        client.verify_authentication()
        print("MAPILLARY_AUTHENTICATION_SUCCESS=True", flush=True)
        retrieval_timestamp = datetime.now(UTC).replace(microsecond=0).isoformat()
        search_boxes = _search_boxes()
        print(f"METADATA_SCAN_BOXES={len(search_boxes)}", flush=True)
        features, query_errors, scanned_boxes = _discover_target_features(
            client, search_boxes
        )
        associations = _feature_image_associations(client, features)
        image_ids = sorted({image_id for _, image_id in associations})
        image_metadata = _fetch_image_metadata(client, image_ids)
        candidate_rows = _build_candidate_rows(
            features,
            associations,
            image_metadata,
            retrieval_timestamp,
        )
        group_map = assign_conservative_independence_groups(candidate_rows)
        for row in candidate_rows:
            row["independence_group_id"] = group_map[row["map_feature_id"]]
        candidate_rows = _deduplicate_candidate_images(candidate_rows)
        summary_rows, dependency_rows = _summaries(
            features,
            associations,
            candidate_rows,
            query_errors,
            scanned_boxes,
        )
        _write_outputs(
            candidate_rows,
            summary_rows,
            dependency_rows,
            client.request_count,
            retrieval_timestamp,
        )
        result = {
            "authentication_success": True,
            "token_serialized_or_logged": False,
            "image_pixels_downloaded": 0,
            "api_metadata_requests": client.request_count,
            "search_boxes_scanned": scanned_boxes,
            "search_box_errors": query_errors,
            "taxonomy_used": {
                label: sorted(
                    {
                        row["exact_mapillary_taxonomy_label"]
                        for row in candidate_rows
                        if row["project_class"] == label
                    }
                )
                for label in TARGET_ORDER
            },
            "summary": summary_rows,
            "artifacts": {
                path.name: {"path": str(path), "sha256": _sha256(path)}
                for path in (PDF_PATH, XLSX_PATH, CSV_PATH)
            },
            "training_prototype_calibration_or_evaluation_performed": False,
        }
        print(json.dumps(result, indent=2))
        return 0
    except (KeyError, MapillaryMetadataError, OSError, TypeError, ValueError) as error:
        print(f"FATAL: {error}", file=sys.stderr)
        return 2


def _ensure_unique_outputs() -> None:
    for path in (PDF_PATH, XLSX_PATH, CSV_PATH):
        if path.exists():
            raise MapillaryMetadataError(
                f"Unique output filename already exists: {path.name}"
            )


def _search_boxes() -> list[dict[str, Any]]:
    boxes: list[dict[str, Any]] = []
    half_span = 0.005
    for city, latitude, longitude in CITY_CENTRES:
        boxes.append(
            {
                "city": city,
                "bbox": (
                    longitude - half_span,
                    latitude - half_span,
                    longitude + half_span,
                    latitude + half_span,
                ),
            }
        )
    return boxes


def _discover_target_features(
    client: MapillaryClient, boxes: Sequence[Mapping[str, Any]]
) -> tuple[dict[str, dict[str, Any]], int, int]:
    features: dict[str, dict[str, Any]] = {}
    errors = 0

    def fetch_box(
        box: Mapping[str, Any],
    ) -> tuple[Mapping[str, Any], list[dict[str, Any]]]:
        bbox = ",".join(f"{float(value):.6f}" for value in box["bbox"])
        payload = client.get_json(
            "/map_features",
            {
                "fields": "id,object_type,object_value,geometry",
                "bbox": bbox,
                "limit": "2000",
            },
        )
        rows = payload.get("data", [])
        if not isinstance(rows, list):
            raise MapillaryMetadataError("Mapillary feature data is not a list")
        return box, [dict(row) for row in rows if isinstance(row, dict)]

    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = [executor.submit(fetch_box, box) for box in boxes]
        completed = 0
        for future in as_completed(futures):
            completed += 1
            try:
                box, rows = future.result()
            except MapillaryMetadataError:
                errors += 1
                continue
            for row in rows:
                taxonomy = str(row.get("object_value", ""))
                project_class = map_exact_taxonomy_label(taxonomy)
                if project_class is None or row.get("object_type") != "trafficsign":
                    continue
                feature_id = str(row.get("id", ""))
                if not re.fullmatch(r"\d+", feature_id):
                    continue
                coordinates = _coordinates(row.get("geometry"))
                if coordinates is None:
                    continue
                existing = features.get(feature_id)
                candidate = {
                    "map_feature_id": feature_id,
                    "exact_mapillary_taxonomy_label": taxonomy,
                    "project_class": project_class,
                    "longitude": coordinates[0],
                    "latitude": coordinates[1],
                    "city_search_area": str(box["city"]),
                    "geographic_evidence_india": (
                        f"Map feature coordinate returned inside curated {box['city']}, India search box"
                    ),
                }
                if existing is None:
                    features[feature_id] = candidate
            if completed % 25 == 0:
                print(
                    f"METADATA_SCAN_PROGRESS={completed}/{len(boxes)};"
                    f"TARGET_FEATURES={len(features)};ERRORS={errors}",
                    flush=True,
                )
    if errors > max(5, math.ceil(len(boxes) * 0.10)):
        raise MapillaryMetadataError(
            f"Too many Mapillary search-box failures: {errors}/{len(boxes)}"
        )
    return features, errors, len(boxes) - errors


def _feature_image_associations(
    client: MapillaryClient, features: Mapping[str, Mapping[str, Any]]
) -> list[tuple[str, str]]:
    associations: set[tuple[str, str]] = set()

    def fetch(feature_id: str) -> tuple[str, list[str]]:
        payload = client.get_json(
            f"/{feature_id}",
            {"fields": "id,object_value,geometry,images"},
        )
        image_ids = _image_ids_from_images_field(payload.get("images"))
        if not image_ids:
            detections = client.get_json(
                f"/{feature_id}/detections",
                {"fields": "id,image", "limit": "100"},
            )
            image_ids = _image_ids_from_detections(detections.get("data"))
        return feature_id, image_ids

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(fetch, feature_id) for feature_id in features]
        for future in as_completed(futures):
            try:
                feature_id, image_ids = future.result()
            except MapillaryMetadataError:
                continue
            for image_id in image_ids:
                associations.add((feature_id, image_id))
    return sorted(associations)


def _fetch_image_metadata(
    client: MapillaryClient, image_ids: Sequence[str]
) -> dict[str, dict[str, Any]]:
    metadata: dict[str, dict[str, Any]] = {}

    def fetch(image_id: str) -> tuple[str, dict[str, Any]]:
        payload = client.get_json(
            f"/{image_id}",
            {"fields": "id,sequence,captured_at,creator,geometry"},
        )
        return image_id, payload

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(fetch, image_id) for image_id in image_ids]
        for future in as_completed(futures):
            try:
                image_id, payload = future.result()
            except MapillaryMetadataError:
                continue
            metadata[image_id] = payload
    return metadata


def _build_candidate_rows(
    features: Mapping[str, Mapping[str, Any]],
    associations: Sequence[tuple[str, str]],
    image_metadata: Mapping[str, Mapping[str, Any]],
    retrieval_timestamp: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for feature_id, image_id in associations:
        feature = features[feature_id]
        image = image_metadata.get(image_id, {})
        coordinates = _coordinates(image.get("geometry")) or (
            float(feature["longitude"]),
            float(feature["latitude"]),
        )
        sequence_id = _nested_id(image.get("sequence"))
        contributor_id, contributor_name = _creator_values(image.get("creator"))
        captured_at = image.get("captured_at")
        captured_epoch = int(str(captured_at)) if captured_at not in (None, "") else ""
        captured_iso = (
            datetime.fromtimestamp(int(str(captured_at)) / 1000.0, UTC)
            .replace(microsecond=0)
            .isoformat()
            if captured_at not in (None, "")
            else ""
        )
        rows.append(
            {
                "mapillary_image_id": image_id,
                "map_feature_id": feature_id,
                "all_linked_map_feature_ids": feature_id,
                "exact_mapillary_taxonomy_label": feature[
                    "exact_mapillary_taxonomy_label"
                ],
                "project_class": feature["project_class"],
                "sequence_id": sequence_id,
                "capture_timestamp_utc": captured_iso,
                "captured_at_epoch_ms": captured_epoch,
                "contributor_id": contributor_id,
                "contributor_name": contributor_name,
                "latitude": f"{coordinates[1]:.7f}",
                "longitude": f"{coordinates[0]:.7f}",
                "geographic_evidence_india": feature["geographic_evidence_india"],
                "city_search_area": feature["city_search_area"],
                "source_reference": f"https://www.mapillary.com/app/?pKey={image_id}",
                "api_retrieval_timestamp_utc": retrieval_timestamp,
                "taxonomy_mapping_confidence": "high_exact_semantic_taxonomy_prefix",
                "independence_group_id": "",
                "review_status": "metadata_only_not_reviewed",
                "approved_photograph": "no",
            }
        )
    return rows


def _deduplicate_candidate_images(
    rows: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    by_key: dict[tuple[str, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        by_key[(str(row["project_class"]), str(row["mapillary_image_id"]))].append(row)
    result: list[dict[str, Any]] = []
    for key in sorted(by_key):
        linked = sorted(by_key[key], key=lambda row: str(row["map_feature_id"]))
        row = dict(linked[0])
        row["all_linked_map_feature_ids"] = ";".join(
            sorted({str(item["map_feature_id"]) for item in linked})
        )
        row["independence_group_id"] = min(
            str(item["independence_group_id"]) for item in linked
        )
        result.append(row)
    return result


def _summaries(
    features: Mapping[str, Mapping[str, Any]],
    associations: Sequence[tuple[str, str]],
    candidates: Sequence[Mapping[str, Any]],
    query_errors: int,
    scanned_boxes: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    feature_class = {
        feature_id: str(feature["project_class"])
        for feature_id, feature in features.items()
    }
    raw_counts = Counter(feature_class[feature_id] for feature_id, _ in associations)
    retained_counts = Counter(str(row["project_class"]) for row in candidates)
    groups: dict[str, set[str]] = defaultdict(set)
    for row in candidates:
        groups[str(row["project_class"])].add(str(row["independence_group_id"]))
    summary: list[dict[str, Any]] = []
    for class_name in TARGET_ORDER:
        local_photos, local_groups = LOCAL_STATE[class_name]
        remaining_photos = max(0, 30 - local_photos)
        remaining_groups = max(0, 15 - local_groups)
        retained = retained_counts[class_name]
        group_count = len(groups[class_name])
        ready = retained >= remaining_photos and group_count >= remaining_groups
        taxonomy = sorted(
            {
                str(row["exact_mapillary_taxonomy_label"])
                for row in candidates
                if row["project_class"] == class_name
            }
        )
        summary.append(
            {
                "project_class": class_name,
                "exact_taxonomy_labels_observed": ";".join(taxonomy),
                "taxonomy_query_rule": EXACT_TAXONOMY_PREFIXES[class_name] + "*",
                "raw_feature_image_associations": raw_counts[class_name],
                "retained_unique_image_candidates": retained,
                "conservative_independent_groups": group_count,
                "dependency_or_duplicate_removals": raw_counts[class_name] - retained,
                "current_local_accepted_photos": local_photos,
                "current_local_accepted_groups": local_groups,
                "remaining_photo_target_before_review": remaining_photos,
                "remaining_group_target_before_review": remaining_groups,
                "metadata_likely_sufficient_for_pixel_review": (
                    "yes" if ready else "no"
                ),
                "metadata_candidates_counted_as_approved": "no",
                "search_scope": (
                    f"{scanned_boxes} curated urban India boxes; {query_errors} box errors"
                ),
            }
        )

    group_rows: list[dict[str, Any]] = []
    by_group: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in candidates:
        by_group[str(row["independence_group_id"])].append(row)
    for group_id in sorted(by_group):
        items = by_group[group_id]
        group_rows.append(
            {
                "independence_group_id": group_id,
                "project_class": items[0]["project_class"],
                "candidate_image_count": len(items),
                "map_feature_ids": ";".join(
                    sorted(
                        {
                            feature_id
                            for item in items
                            for feature_id in str(
                                item["all_linked_map_feature_ids"]
                            ).split(";")
                        }
                    )
                ),
                "sequence_ids": ";".join(
                    sorted(
                        {
                            str(item["sequence_id"])
                            for item in items
                            if item["sequence_id"]
                        }
                    )
                ),
                "contributor_ids": ";".join(
                    sorted(
                        {
                            str(item["contributor_id"])
                            for item in items
                            if item["contributor_id"]
                        }
                    )
                ),
                "cities": ";".join(
                    sorted({str(item["city_search_area"]) for item in items})
                ),
                "grouping_rule": (
                    "same feature; same image-linked features; <=20m same-class; or <=75m with same sequence/contributor and <=120s"
                ),
            }
        )
    return summary, group_rows


def _write_outputs(
    candidates: list[dict[str, Any]],
    summary: list[dict[str, Any]],
    dependency_groups: list[dict[str, Any]],
    request_count: int,
    retrieval_timestamp: str,
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(
        prefix="mapillary-india-metadata-", dir=OUTPUT_DIR
    ) as temporary_directory:
        stage = Path(temporary_directory)
        staged_pdf = stage / PDF_PATH.name
        staged_xlsx = stage / XLSX_PATH.name
        staged_csv = stage / CSV_PATH.name
        _write_csv(staged_csv, candidates)
        _write_workbook(staged_xlsx, summary, candidates, dependency_groups)
        _write_pdf(
            staged_pdf,
            summary,
            dependency_groups,
            request_count,
            retrieval_timestamp,
        )
        _validate_outputs(
            staged_pdf,
            staged_xlsx,
            staged_csv,
            summary,
            candidates,
            dependency_groups,
        )
        os.replace(staged_pdf, PDF_PATH)
        os.replace(staged_xlsx, XLSX_PATH)
        os.replace(staged_csv, CSV_PATH)


def _write_csv(path: Path, candidates: Sequence[Mapping[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=CANDIDATE_COLUMNS, extrasaction="raise"
        )
        writer.writeheader()
        writer.writerows(candidates)


def _write_workbook(
    path: Path,
    summary: list[dict[str, Any]],
    candidates: list[dict[str, Any]],
    dependency_groups: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_sheet(workbook, "Summary", summary)
    for class_name in TARGET_ORDER:
        rows = [row for row in candidates if row["project_class"] == class_name]
        _add_sheet(
            workbook,
            SHEET_NAMES[class_name],
            rows,
            fallback_columns=CANDIDATE_COLUMNS,
        )
    _add_sheet(
        workbook,
        "Dependency Groups",
        dependency_groups,
        fallback_columns=(
            "independence_group_id",
            "project_class",
            "candidate_image_count",
            "map_feature_ids",
            "sequence_ids",
            "contributor_ids",
            "cities",
            "grouping_rule",
        ),
    )
    _add_sheet(workbook, "Terms Snapshot", list(TERMS_ROWS))
    workbook.save(path)


def _add_sheet(
    workbook: Workbook,
    name: str,
    rows: Sequence[Mapping[str, Any]],
    fallback_columns: Sequence[str] | None = None,
) -> None:
    sheet = workbook.create_sheet(name)
    columns = tuple(rows[0]) if rows else tuple(fallback_columns or ())
    if not columns:
        raise MapillaryMetadataError(f"Cannot infer columns for sheet {name}")
    sheet.append(list(columns))
    for row in rows:
        sheet.append([row[column] for column in columns])
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, 1):
        values = [str(column), *(str(row[column]) for row in rows)]
        width = min(max(len(value) for value in values) + 2, 55)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 12)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_pdf(
    path: Path,
    summary: list[dict[str, Any]],
    dependency_groups: list[dict[str, Any]],
    request_count: int,
    retrieval_timestamp: str,
) -> None:
    styles = getSampleStyleSheet()
    small = ParagraphStyle(
        "MapillarySmall",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        spaceAfter=2 * mm,
    )
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=11 * mm,
        rightMargin=11 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title="Mapillary India metadata discovery report",
        author="Adaptive Indian Road Sign Recognition project",
    )
    story: list[Any] = [
        Paragraph("Mapillary India Metadata Discovery", styles["Title"]),
        Paragraph(
            f"Authenticated metadata retrieval completed at {_escape(retrieval_timestamp)}. API metadata requests: {request_count}. Image pixels downloaded: 0. No candidate is visually reviewed or approved.",
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
        _rows_table(
            summary,
            (
                "project_class",
                "exact_taxonomy_labels_observed",
                "raw_feature_image_associations",
                "retained_unique_image_candidates",
                "conservative_independent_groups",
                "dependency_or_duplicate_removals",
                "metadata_likely_sufficient_for_pixel_review",
            ),
            [45 * mm, 75 * mm, 35 * mm, 38 * mm, 38 * mm, 38 * mm, 48 * mm],
            7,
        ),
        PageBreak(),
        Paragraph("Authentication and exact taxonomy", styles["Heading1"]),
        Paragraph(
            "Authentication succeeded using MAPILLARY_ACCESS_TOKEN in an HTTP authorization header. The token was never placed in a URL, report, output file, or log. Only taxonomy labels beginning with the following exact semantic prefixes were retained:",
            styles["BodyText"],
        ),
        _table(
            [["Project class", "Accepted Mapillary taxonomy rule"]]
            + [
                [label, prefix + "*"]
                for label, prefix in EXACT_TAXONOMY_PREFIXES.items()
            ],
            [85 * mm, 185 * mm],
            9,
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "Labels that were merely similar, generic, or missing the exact speed numeral/turn direction were excluded. Metadata candidates remain unapproved photographs.",
            styles["BodyText"],
        ),
        PageBreak(),
        Paragraph("India-only geographic validation", styles["Heading1"]),
        Paragraph(
            "The scan used one approximately 1.0-km-wide bounding box centred on each of 45 named Indian urban areas. Every retained feature coordinate came from one of these boxes and carries the city evidence in the workbook/CSV. No border-crossing or uncertain-country box was queried. This is a conservative city-centre discovery sample, not an exhaustive census of all Mapillary coverage in India.",
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
        _rows_table(
            summary,
            (
                "project_class",
                "search_scope",
                "retained_unique_image_candidates",
                "conservative_independent_groups",
            ),
            [60 * mm, 105 * mm, 55 * mm, 55 * mm],
            8,
        ),
        PageBreak(),
        Paragraph("Conservative dependency methodology", styles["Heading1"]),
        Paragraph(
            "All images linked to the same Mapillary map-feature ID are one physical-sign group. Features are also merged when the same image links them, when same-class feature coordinates are within 20 metres, or when features are within 75 metres and share sequence/contributor evidence with capture times within 120 seconds. A whole sequence is not collapsed automatically because one drive can contain multiple physical signs.",
            styles["BodyText"],
        ),
        Spacer(1, 4 * mm),
        _table(
            [
                ["Measure", "Value"],
                ["Dependency groups", str(len(dependency_groups))],
                ["Pixel inspection", "not performed"],
                ["Human approval", "not performed"],
                ["Experiment use", "blocked"],
            ],
            [70 * mm, 180 * mm],
            9,
        ),
        PageBreak(),
        Paragraph("Quantity analysis and readiness", styles["Heading1"]),
        _rows_table(
            summary,
            (
                "project_class",
                "current_local_accepted_photos",
                "current_local_accepted_groups",
                "remaining_photo_target_before_review",
                "remaining_group_target_before_review",
                "retained_unique_image_candidates",
                "conservative_independent_groups",
                "metadata_likely_sufficient_for_pixel_review",
            ),
            [45 * mm, 30 * mm, 30 * mm, 35 * mm, 35 * mm, 38 * mm, 38 * mm, 45 * mm],
            7,
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "Readiness means only that metadata volume could justify a later pixel-review request. It does not add Mapillary records to the approved-photo totals. The locked requirement remains 30 approved photographs and 15 independent groups per class.",
            styles["BodyText"],
        ),
        PageBreak(),
        Paragraph("Terms and attribution snapshot", styles["Heading1"]),
        _rows_table(
            list(TERMS_ROWS),
            ("topic", "finding", "reference", "retrieval_date", "project_implication"),
            [35 * mm, 80 * mm, 75 * mm, 28 * mm, 75 * mm],
            7,
        ),
        Spacer(1, 4 * mm),
        Paragraph(
            "No claim is made here that private-repository storage, derived-crop redistribution, or report-image publication is automatically permitted beyond the cited guidance. Those questions require a deliberate logged-in Terms review and an attribution plan before pixels are downloaded.",
            small,
        ),
        PageBreak(),
        Paragraph("Exact next action", styles["Heading1"]),
        Paragraph(
            "Review the per-class metadata tables and dependency groups. For classes marked metadata-sufficient, prepare a separate, user-approved pixel-acquisition manifest containing only selected image IDs, required attribution fields, and one or more candidates per conservative group. Before executing it, verify the current logged-in Terms for academic use, report examples, derived crops, and private repository storage. Stop remains in effect: do not download pixels, visually approve candidates, create prototypes, train, calibrate, or evaluate in this phase.",
            styles["BodyText"],
        ),
    ]
    document.build(story)


def _validate_outputs(
    pdf_path: Path,
    xlsx_path: Path,
    csv_path: Path,
    summary: Sequence[Mapping[str, Any]],
    candidates: Sequence[Mapping[str, Any]],
    dependency_groups: Sequence[Mapping[str, Any]],
) -> None:
    pages = len(PdfReader(str(pdf_path)).pages)
    if pages != 7:
        raise MapillaryMetadataError(f"Expected 7 PDF pages, found {pages}")
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    expected_sheets = [
        "Summary",
        "STOP",
        "Speed 50",
        "No Left Turn",
        "No Parking",
        "Dependency Groups",
        "Terms Snapshot",
    ]
    if workbook.sheetnames != expected_sheets:
        workbook.close()
        raise MapillaryMetadataError(
            f"Unexpected workbook sheets: {workbook.sheetnames}"
        )
    if workbook["Summary"].max_row != 5:
        workbook.close()
        raise MapillaryMetadataError("Summary sheet must contain four classes")
    if workbook["Dependency Groups"].max_row - 1 != len(dependency_groups):
        workbook.close()
        raise MapillaryMetadataError("Dependency group row count mismatch")
    workbook.close()
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        csv_rows = [dict(row) for row in csv.DictReader(handle)]
    if len(csv_rows) != len(candidates):
        raise MapillaryMetadataError("CSV candidate row count mismatch")
    if any(row["approved_photograph"] != "no" for row in csv_rows):
        raise MapillaryMetadataError("Metadata candidate was marked approved")
    if any(
        map_exact_taxonomy_label(row["exact_mapillary_taxonomy_label"])
        != row["project_class"]
        for row in csv_rows
    ):
        raise MapillaryMetadataError("CSV contains an approximate taxonomy mapping")
    if len(summary) != 4:
        raise MapillaryMetadataError("Summary must contain exactly four target classes")


def _image_ids_from_images_field(value: Any) -> list[str]:
    if not isinstance(value, dict) or not isinstance(value.get("data"), list):
        return []
    return sorted(
        {
            str(item["id"])
            for item in value["data"]
            if isinstance(item, dict) and re.fullmatch(r"\d+", str(item.get("id", "")))
        }
    )


def _image_ids_from_detections(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    result: set[str] = set()
    for item in value:
        if not isinstance(item, dict):
            continue
        image = item.get("image")
        image_id = _nested_id(image)
        if re.fullmatch(r"\d+", image_id):
            result.add(image_id)
    return sorted(result)


def _coordinates(value: Any) -> tuple[float, float] | None:
    if not isinstance(value, dict):
        return None
    coordinates = value.get("coordinates")
    if (
        not isinstance(coordinates, list)
        or len(coordinates) < 2
        or not all(isinstance(item, (int, float)) for item in coordinates[:2])
    ):
        return None
    return float(coordinates[0]), float(coordinates[1])


def _nested_id(value: Any) -> str:
    if isinstance(value, dict):
        return str(value.get("id", ""))
    if value in (None, ""):
        return ""
    return str(value)


def _creator_values(value: Any) -> tuple[str, str]:
    if not isinstance(value, dict):
        return "", ""
    return str(value.get("id", "")), str(value.get("username", value.get("name", "")))


def _rows_table(
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[str],
    widths: Sequence[float],
    font_size: float,
) -> Table:
    data = [[_header(column) for column in columns]]
    data.extend([[str(row[column]) for column in columns] for row in rows])
    return _table(data, widths, font_size)


def _table(
    data: Sequence[Sequence[Any]], widths: Sequence[float], font_size: float
) -> Table:
    body = ParagraphStyle(
        f"Body{font_size}",
        fontName="Helvetica",
        fontSize=font_size,
        leading=font_size + 2,
    )
    header = ParagraphStyle(
        f"Header{font_size}",
        parent=body,
        fontName="Helvetica-Bold",
        textColor=colors.white,
    )
    formatted = [
        [
            Paragraph(_escape(str(value)), header if row_index == 0 else body)
            for value in row
        ]
        for row_index, row in enumerate(data)
    ]
    table = Table(formatted, colWidths=list(widths), repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#EDF3F8")],
                ),
            ]
        )
    )
    return table


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _escape(value: str) -> str:
    return value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _header(value: str) -> str:
    return value.replace("_", " ").title()


if __name__ == "__main__":
    raise SystemExit(main())
