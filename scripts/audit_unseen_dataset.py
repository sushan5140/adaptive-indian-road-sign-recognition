"""Audit quarantined unseen-sign photographs and prepare review artifacts."""

from __future__ import annotations

import argparse
import csv
import json
import logging
import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from data.unseen_audit import (  # noqa: E402
    REVIEW_COLUMNS,
    UnseenAuditConfig,
    UnseenAuditError,
    UnseenDatasetAuditor,
)

LOGGER = logging.getLogger(__name__)


def build_parser() -> argparse.ArgumentParser:
    """Build command-line arguments for unseen-class intake auditing."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", default="configs/open_set_unseen.yaml")
    return parser


def main(argv: list[str] | None = None) -> int:
    """Run the audit, preserving validated decisions, and write review files."""
    logging.basicConfig(
        level=logging.INFO,
        format=(
            '{"time":"%(asctime)s","level":"%(levelname)s",'
            '"logger":"%(name)s","message":"%(message)s"}'
        ),
    )
    arguments = build_parser().parse_args(argv)
    try:
        config_payload = _read_yaml(_project_path(arguments.config))
        paths = _mapping(config_payload, "paths")
        audit_values = _mapping(config_payload, "audit")
        candidates = _candidate_rows(config_payload)
        class_names = tuple(str(row["class_name"]) for row in candidates)
        review_csv = _project_path(str(paths["review_csv"]))
        review_xlsx = _project_path(str(paths["review_xlsx"]))
        output = _project_path(str(paths["audit_output_dir"]))
        result = UnseenDatasetAuditor(
            UnseenAuditConfig(
                raw_root=_project_path(str(paths["raw_root"])),
                class_names=class_names,
                source_metadata_csv=_project_path(str(paths["source_metadata_csv"])),
                existing_review_csv=review_csv,
                allowed_extensions=tuple(
                    str(value) for value in audit_values["allowed_extensions"]
                ),
                near_duplicate_hamming_distance=int(
                    audit_values["near_duplicate_hamming_distance"]
                ),
                minimum_independent_groups=int(
                    audit_values["minimum_independent_groups"]
                ),
                target_photos_minimum=int(audit_values["target_photos_minimum"]),
                target_photos_maximum=int(audit_values["target_photos_maximum"]),
            )
        ).audit()
        candidate_report = _enrich_candidates(candidates, result.summary)
        output.mkdir(parents=True, exist_ok=True)
        _write_csv(review_csv, list(result.review_rows), REVIEW_COLUMNS)
        _write_workbook(
            review_xlsx, "Unseen Review", REVIEW_COLUMNS, result.review_rows
        )
        _write_csv(
            output / "unseen_review.csv", list(result.review_rows), REVIEW_COLUMNS
        )
        _write_workbook(
            output / "unseen_review.xlsx",
            "Unseen Review",
            REVIEW_COLUMNS,
            result.review_rows,
        )
        _write_json(output / "unseen_intake_audit_summary.json", result.summary)
        _write_csv(
            output / "near_duplicate_pairs.csv",
            list(result.near_duplicate_pairs),
            (
                "left_review_id",
                "right_review_id",
                "hamming_distance",
                "cross_label",
            ),
        )
        candidate_columns = tuple(candidate_report[0])
        _write_csv(
            output / "candidate_unseen_classes.csv",
            candidate_report,
            candidate_columns,
        )
        _write_workbook(
            output / "candidate_unseen_classes.xlsx",
            "Candidate Classes",
            candidate_columns,
            candidate_report,
        )
    except (OSError, KeyError, TypeError, ValueError, UnseenAuditError) as error:
        LOGGER.error("Unseen-class audit failed: %s", error)
        return 2
    LOGGER.info("Candidate classes: %d", len(candidate_report))
    LOGGER.info("Raw photographs: %d", len(result.review_rows))
    LOGGER.info(
        "Pending review: %d", result.summary["review_status_counts"].get("pending", 0)
    )
    LOGGER.info("Partitions created: no")
    LOGGER.info("Training or evaluation performed: no")
    return 0


def _candidate_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    values = payload.get("candidate_classes")
    if not isinstance(values, list) or not values:
        raise UnseenAuditError("candidate_classes must be a non-empty list")
    rows: list[dict[str, Any]] = []
    for value in values:
        if not isinstance(value, dict):
            raise UnseenAuditError("Each candidate class must be a mapping")
        rows.append(value)
    return rows


def _enrich_candidates(
    candidates: list[dict[str, Any]], summary: dict[str, Any]
) -> list[dict[str, Any]]:
    minimum = int(summary["minimum_independent_groups_per_class"])
    target = summary["target_photos_per_class"]
    per_class = summary["per_class"]
    rows: list[dict[str, Any]] = []
    for candidate in candidates:
        class_name = str(candidate["class_name"])
        measured = per_class[class_name]
        rows.append(
            {
                "class_name": class_name,
                "visual_role": str(candidate["visual_role"]),
                "nearest_base_classes": ";".join(candidate["nearest_base_classes"]),
                "rationale": str(candidate["rationale"]),
                "local_data_warning": str(candidate["local_data_warning"]),
                "operational_minimum_independent_groups": minimum,
                "target_photos": f"{target[0]}-{target[1]}",
                "current_raw_photos": measured["raw_photos"],
                "current_independent_groups": measured["independent_groups"],
                "current_pending": measured["pending"],
                "enough_approved_data_to_partition": str(
                    measured["minimum_groups_met"]
                ).lower(),
            }
        )
    return rows


def _read_yaml(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        payload = yaml.safe_load(handle)
    if not isinstance(payload, dict):
        raise UnseenAuditError(f"Configuration root must be a mapping: {path}")
    return payload


def _mapping(payload: dict[str, Any], key: str) -> dict[str, Any]:
    value = payload.get(key)
    if not isinstance(value, dict):
        raise UnseenAuditError(f"Configuration section {key!r} must be a mapping")
    return value


def _write_csv(
    path: Path, rows: list[dict[str, Any]], columns: tuple[str, ...]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )


def _write_workbook(
    path: Path,
    sheet_name: str,
    columns: tuple[str, ...],
    rows: tuple[dict[str, str], ...] | list[dict[str, Any]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(list(columns))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, start=1):
        measured = max(
            [len(column)]
            + [
                len(str(sheet.cell(row=row, column=index).value or ""))
                for row in range(2, sheet.max_row + 1)
            ]
        )
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = (
            min(measured + 2, 60)
        )
    workbook.save(path)


def _project_path(value: str) -> Path:
    path = Path(value).expanduser()
    return path.resolve() if path.is_absolute() else (PROJECT_ROOT / path).resolve()


if __name__ == "__main__":
    raise SystemExit(main())
