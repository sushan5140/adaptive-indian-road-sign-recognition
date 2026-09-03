"""Apply a completed local unseen-class review as a protected decision overlay."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import sys
import tempfile
from collections import Counter, defaultdict
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from data.local_unseen_review import (  # noqa: E402
    DECISION_FIELDS,
    REVIEW_COLUMNS,
    LocalUnseenReviewError,
    is_experiment_ready,
    validate_human_review_overlay,
)

COLUMNS = REVIEW_COLUMNS

BASE_PATH = PROJECT_ROOT / "outputs/open_set_acquisition/local_unseen_review.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "outputs/open_set_acquisition"
OUTPUT_STEM = "unseen43_human_review_applied_20260901_v1"
XLSX_PATH = OUTPUT_DIR / f"{OUTPUT_STEM}.xlsx"
CSV_PATH = OUTPUT_DIR / f"{OUTPUT_STEM}.csv"
SUMMARY_PATH = OUTPUT_DIR / f"{OUTPUT_STEM}_summary.json"
MINIMUM_GROUPS = 15
TARGET_PHOTOS_MINIMUM = 30
ACTIVE_CLASSES = (
    "stop",
    "no_left_turn",
    "maximum_speed_limit_50_km_h",
    "no_parking",
    "bus_stop",
)


def build_parser() -> argparse.ArgumentParser:
    """Build command-line arguments for the protected review import."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "reviewed_workbook",
        type=Path,
        help="Completed 43-row workbook whose decision fields should be imported.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    """Validate the review overlay and write licence-blocked reviewed artifacts."""
    arguments = build_parser().parse_args(argv)
    reviewed_path = arguments.reviewed_workbook.resolve()
    try:
        base_rows = _read_workbook_rows(BASE_PATH, expected_sheets=("Pending Review",))
        reviewed_rows = _read_workbook_rows(
            reviewed_path,
            expected_sheets=("Pending Review", "Human Review Summary"),
        )
        class_status = validate_human_review_overlay(base_rows, reviewed_rows)
        imported_rows = _apply_decisions(base_rows, reviewed_rows)
        validate_human_review_overlay(base_rows, imported_rows)
        summary = _build_summary(imported_rows, class_status, reviewed_path)

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(
            prefix="apply-local-unseen-review-", dir=OUTPUT_DIR
        ) as temporary_directory:
            stage = Path(temporary_directory)
            staged_xlsx = stage / XLSX_PATH.name
            staged_csv = stage / CSV_PATH.name
            staged_summary = stage / SUMMARY_PATH.name
            _write_csv(staged_csv, imported_rows)
            _write_workbook(staged_xlsx, imported_rows, summary)
            _write_json(staged_summary, summary)
            _validate_outputs(
                staged_xlsx,
                staged_csv,
                staged_summary,
                base_rows,
                imported_rows,
            )
            os.replace(staged_xlsx, XLSX_PATH)
            os.replace(staged_csv, CSV_PATH)
            os.replace(staged_summary, SUMMARY_PATH)

        result = {
            **summary,
            "artifacts": {
                path.name: {"path": str(path), "sha256": _sha256(path)}
                for path in (XLSX_PATH, CSV_PATH, SUMMARY_PATH)
            },
        }
        print(json.dumps(result, indent=2))
        return 0
    except (KeyError, LocalUnseenReviewError, OSError, TypeError, ValueError) as error:
        print(f"FATAL: {error}", file=sys.stderr)
        return 2


def _apply_decisions(
    base_rows: Sequence[Mapping[str, str]],
    reviewed_rows: Sequence[Mapping[str, str]],
) -> list[dict[str, str]]:
    reviewed_by_id = {row["review_id"]: row for row in reviewed_rows}
    imported: list[dict[str, str]] = []
    for base in base_rows:
        row = dict(base)
        decision = reviewed_by_id[row["review_id"]]
        for field in DECISION_FIELDS:
            row[field] = decision[field]
        imported.append(row)
    return imported


def _build_summary(
    rows: Sequence[Mapping[str, str]],
    class_status: Mapping[str, Mapping[str, int]],
    reviewed_path: Path,
) -> dict[str, Any]:
    accepted = [row for row in rows if row["review_status"] == "accepted"]
    rejected = [row for row in rows if row["review_status"] == "rejected"]
    groups: dict[str, set[str]] = defaultdict(set)
    accepted_images = Counter[str]()
    for row in accepted:
        label = row["proposed_unseen_class"]
        accepted_images[label] += 1
        groups[label].add(row["source_or_independence_group"])

    acquisition_routes = {
        "stop": "Acquire licensed Indian real photographs through a metadata-first Mapillary/IRSDB search; retain a cross-domain GTSRB fallback only if documented.",
        "maximum_speed_limit_50_km_h": "Acquire licensed Indian speed-50 photographs; the two local annotations remain visually invalid and excluded.",
        "no_left_turn": "Acquire additional licensed independent photographs because one accepted group is scientifically insufficient.",
        "no_parking": "Acquire additional licensed independent photographs because three accepted groups are scientifically insufficient.",
    }
    gaps: list[dict[str, Any]] = []
    for label in (
        "stop",
        "maximum_speed_limit_50_km_h",
        "no_left_turn",
        "no_parking",
    ):
        image_count = accepted_images[label]
        group_count = len(groups[label])
        gaps.append(
            {
                "class_name": label,
                "visually_accepted_images": image_count,
                "visually_accepted_independent_groups": group_count,
                "minimum_independent_groups": MINIMUM_GROUPS,
                "additional_groups_needed": max(0, MINIMUM_GROUPS - group_count),
                "target_photos_minimum": TARGET_PHOTOS_MINIMUM,
                "additional_photos_needed_to_target_minimum": max(
                    0, TARGET_PHOTOS_MINIMUM - image_count
                ),
                "licence_ready_images": 0,
                "recommended_route": acquisition_routes[label],
            }
        )

    experiment_ready = [row for row in rows if is_experiment_ready(row)]
    return {
        "active_unseen_classes": list(ACTIVE_CLASSES),
        "authoritative_input": str(reviewed_path),
        "authoritative_input_sha256": _sha256(reviewed_path),
        "base_pending_manifest": str(BASE_PATH),
        "base_pending_manifest_sha256": _sha256(BASE_PATH),
        "review_rows": len(rows),
        "class_status_counts": {
            label: dict(counts) for label, counts in class_status.items()
        },
        "total_status_counts": dict(Counter(row["review_status"] for row in rows)),
        "visually_accepted_but_licence_blocked_count": len(accepted),
        "visually_accepted_but_licence_blocked_ids": [
            row["review_id"] for row in accepted
        ],
        "rejected_ids": [row["review_id"] for row in rejected],
        "licence_status_counts": dict(Counter(row["licence_status"] for row in rows)),
        "accepted_independent_group_counts": {
            label: len(groups[label])
            for label in ("bus_stop", "no_left_turn", "no_parking")
        },
        "experiment_ready_count": len(experiment_ready),
        "experiment_readiness_rule": (
            "review_status must equal accepted AND licence_status must equal "
            "approved or confirmed"
        ),
        "acquisition_thresholds_unchanged": {
            "minimum_independent_groups": MINIMUM_GROUPS,
            "target_photos_minimum": TARGET_PHOTOS_MINIMUM,
        },
        "remaining_acquisition_gaps": gaps,
        "protected_field_differences": 0,
        "frozen_v2_modified": False,
        "training_prototype_calibration_or_evaluation_performed": False,
    }


def _read_workbook_rows(
    path: Path, *, expected_sheets: tuple[str, ...]
) -> list[dict[str, str]]:
    if not path.is_file():
        raise LocalUnseenReviewError(f"Workbook does not exist: {path}")
    workbook = load_workbook(path, read_only=True, data_only=False)
    if tuple(workbook.sheetnames) != expected_sheets:
        workbook.close()
        raise LocalUnseenReviewError(
            f"Unexpected workbook sheets in {path.name}: {workbook.sheetnames}"
        )
    sheet = workbook["Pending Review"]
    values = sheet.iter_rows(values_only=True)
    headers = tuple(str(value) for value in next(values))
    if headers != COLUMNS:
        workbook.close()
        raise LocalUnseenReviewError(
            f"Workbook schema differs from the locked schema: {path}"
        )
    rows: list[dict[str, str]] = []
    for values_row in values:
        if not any(value is not None for value in values_row):
            continue
        if any(
            isinstance(value, str) and value.startswith("=") for value in values_row
        ):
            workbook.close()
            raise LocalUnseenReviewError(f"Formulas are not permitted: {path}")
        rows.append(
            {
                header: "" if value is None else str(value)
                for header, value in zip(headers, values_row, strict=True)
            }
        )
    workbook.close()
    return rows


def _write_csv(path: Path, rows: Sequence[Mapping[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS, extrasaction="raise")
        writer.writeheader()
        writer.writerows(rows)


def _write_workbook(
    path: Path, rows: Sequence[Mapping[str, str]], summary: Mapping[str, Any]
) -> None:
    workbook = Workbook()
    review_sheet = workbook.active
    review_sheet.title = "Reviewed Candidates"
    _write_sheet(
        review_sheet, COLUMNS, [[row[column] for column in COLUMNS] for row in rows]
    )

    summary_sheet = workbook.create_sheet("Decision Summary")
    summary_columns = ("proposed_unseen_class", "accepted", "rejected")
    summary_rows = [
        [label, counts["accepted"], counts["rejected"]]
        for label, counts in summary["class_status_counts"].items()
    ]
    _write_sheet(summary_sheet, summary_columns, summary_rows)

    gaps_sheet = workbook.create_sheet("Acquisition Gaps")
    gaps = summary["remaining_acquisition_gaps"]
    gap_columns = tuple(gaps[0])
    _write_sheet(
        gaps_sheet,
        gap_columns,
        [[row[column] for column in gap_columns] for row in gaps],
    )
    workbook.save(path)


def _write_sheet(
    sheet: Any, columns: Sequence[str], rows: Sequence[Sequence[Any]]
) -> None:
    sheet.append(list(columns))
    for row in rows:
        sheet.append(list(row))
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, 1):
        values = [str(column), *(str(row[index - 1]) for row in rows)]
        width = min(max(len(value) for value in values) + 2, 60)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 12)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_json(path: Path, payload: Mapping[str, Any]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)
        handle.write("\n")


def _validate_outputs(
    xlsx_path: Path,
    csv_path: Path,
    summary_path: Path,
    base_rows: Sequence[Mapping[str, str]],
    expected_rows: Sequence[Mapping[str, str]],
) -> None:
    with csv_path.open(encoding="utf-8-sig", newline="") as handle:
        csv_rows = [dict(row) for row in csv.DictReader(handle)]
    validate_human_review_overlay(base_rows, csv_rows)
    if csv_rows != list(expected_rows):
        raise LocalUnseenReviewError("Generated CSV differs from imported rows")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    if workbook.sheetnames != [
        "Reviewed Candidates",
        "Decision Summary",
        "Acquisition Gaps",
    ]:
        workbook.close()
        raise LocalUnseenReviewError(
            f"Unexpected generated workbook sheets: {workbook.sheetnames}"
        )
    sheet = workbook["Reviewed Candidates"]
    xlsx_ids = [
        str(sheet.cell(index, 1).value) for index in range(2, sheet.max_row + 1)
    ]
    workbook.close()
    if xlsx_ids != [row["review_id"] for row in expected_rows]:
        raise LocalUnseenReviewError("Generated XLSX review IDs differ from CSV")

    with summary_path.open(encoding="utf-8") as handle:
        summary = json.load(handle)
    if summary["experiment_ready_count"] != 0:
        raise LocalUnseenReviewError(
            "Pending licences must block all candidates from experiment use"
        )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


if __name__ == "__main__":
    raise SystemExit(main())
