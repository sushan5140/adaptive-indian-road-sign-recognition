"""Create the pending-only local Dataset B unseen-class review bundle."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import sys
import tempfile
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PilImage
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from data.local_unseen_review import (  # noqa: E402
    REVIEW_COLUMNS,
    LocalUnseenReviewError,
    validate_local_unseen_review_rows,
)
from data.source_discovery import (  # noqa: E402
    SourceDiscoveryError,
    build_perceptual_groups,
    frozen_split_overlap,
    make_split_evidence,
)

OUTPUT_DIR = PROJECT_ROOT / "outputs/open_set_acquisition"
PDF_PATH = OUTPUT_DIR / "local_unseen_review_bundle.pdf"
XLSX_PATH = OUTPUT_DIR / "local_unseen_review.xlsx"
CSV_PATH = OUTPUT_DIR / "local_unseen_review.csv"
IMAGE_ROOT = PROJECT_ROOT / "data/raw/indian_traffic_vqa/traffic512final"
INVENTORY_PATH = PROJECT_ROOT / "outputs/dataset_b_audit/image_inventory.csv"
NEAR_PAIRS_PATH = PROJECT_ROOT / "outputs/dataset_b_audit/near_duplicate_pairs.csv"
EXCLUDED_PATH = PROJECT_ROOT / "outputs/dataset_b_audit/excluded_candidates.csv"
V2_REVIEW_PATH = PROJECT_ROOT / "outputs/v2_review/dataset_b_v2_review.csv"
CONFIG_PATH = PROJECT_ROOT / "configs/open_set_unseen.yaml"

ACTIVE_UNSEEN_CLASSES = (
    "stop",
    "no_left_turn",
    "maximum_speed_limit_50_km_h",
    "no_parking",
    "bus_stop",
)
BUNDLE_CLASS_ORDER = ("bus_stop", "no_left_turn", "no_parking")
LOCAL_ELIGIBILITY: dict[str, tuple[str, str]] = {
    "img_0579.jpg": (
        "no_left_turn",
        "Clear circular no-left-turn sign; requires a new unseen-class decision.",
    ),
    "img_1021.jpg": (
        "no_left_turn",
        "Distant but visible no-left-turn sign; requires a new unseen-class decision.",
    ),
    "img_0786.jpg": (
        "no_parking",
        "Visible no-parking sign; conservatively grouped with img_0812.jpg.",
    ),
    "img_0812.jpg": (
        "no_parking",
        "Visible no-parking sign; conservatively grouped with img_0786.jpg.",
    ),
    "img_0816.jpg": (
        "no_parking",
        "Visible no-parking sign in a separate roadside scene.",
    ),
    "img_0970.jpg": (
        "no_parking",
        "Police no-parking board; visually eligible but stylistically non-standard.",
    ),
    "img_1016.jpg": (
        "no_parking",
        "Visible no-parking regulatory sign in a distinct scene.",
    ),
}

COLUMNS = REVIEW_COLUMNS


def main() -> int:
    """Build and validate the local unseen-class manual-review package."""
    try:
        _validate_active_protocol()
        rows = _build_rows()
        measured = validate_local_unseen_review_rows(rows)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(
            prefix="local-unseen-review-", dir=OUTPUT_DIR
        ) as temporary_directory:
            stage = Path(temporary_directory)
            staged_pdf = stage / PDF_PATH.name
            staged_xlsx = stage / XLSX_PATH.name
            staged_csv = stage / CSV_PATH.name
            _write_csv(staged_csv, rows)
            _write_workbook(staged_xlsx, rows)
            _write_pdf(staged_pdf, rows)
            page_count = _validate_artifacts(staged_pdf, staged_xlsx, staged_csv, rows)
            os.replace(staged_pdf, PDF_PATH)
            os.replace(staged_xlsx, XLSX_PATH)
            os.replace(staged_csv, CSV_PATH)
        result = {
            "active_unseen_classes": list(ACTIVE_UNSEEN_CLASSES),
            "review_rows": len(rows),
            "per_class": measured,
            "pdf_pages": page_count,
            "xlsx_sheets": ["Pending Review"],
            "licence_status": "pending_curator_confirmation",
            "artifacts": {
                path.name: {"path": str(path), "sha256": _sha256(path)}
                for path in (PDF_PATH, XLSX_PATH, CSV_PATH)
            },
            "training_prototype_calibration_or_evaluation_performed": False,
        }
        print(json.dumps(result, indent=2))
        return 0
    except (
        KeyError,
        LocalUnseenReviewError,
        OSError,
        SourceDiscoveryError,
        TypeError,
        ValueError,
    ) as error:
        print(f"FATAL: {error}", file=sys.stderr)
        return 2


def _validate_active_protocol() -> None:
    with CONFIG_PATH.open(encoding="utf-8") as handle:
        payload = yaml.safe_load(handle)
    configured = tuple(
        str(row["class_name"]) for row in payload.get("candidate_classes", [])
    )
    if configured != ACTIVE_UNSEEN_CLASSES:
        raise LocalUnseenReviewError(
            f"Active unseen protocol mismatch: {configured}; expected {ACTIVE_UNSEEN_CLASSES}"
        )
    if "roundabout_ahead" in configured:
        raise LocalUnseenReviewError("roundabout_ahead must not remain active")


def _build_rows() -> list[dict[str, str]]:
    inventory = {row["image_name"]: row for row in _read_csv(INVENTORY_PATH)}
    image_hashes = {name: row["sha256"] for name, row in inventory.items()}
    groups = build_perceptual_groups(inventory, _read_csv(NEAR_PAIRS_PATH))
    excluded = {row["image_id"]: row for row in _read_csv(EXCLUDED_PATH)}
    v2_review = _read_csv(V2_REVIEW_PATH)
    split_rows = {
        split: _read_csv(PROJECT_ROOT / f"outputs/manifests/v2_{split}.csv")
        for split in ("train", "validation", "test")
    }
    split_evidence = {
        split: make_split_evidence(rows, image_hashes)
        for split, rows in split_rows.items()
    }
    all_split_rows = [row for rows in split_rows.values() for row in rows]
    all_evidence = make_split_evidence(all_split_rows, image_hashes)

    candidates: list[dict[str, str]] = []
    for row in v2_review:
        if row["proposed_class"] != "stop":
            continue
        if row["review_status"] != "rejected":
            raise LocalUnseenReviewError(
                f"Expected rejected V2 stop row: {row['review_id']}"
            )
        candidates.append(
            _make_row(
                source_id=row["source_image_id"],
                new_label="bus_stop",
                source_question=row["source_question"],
                source_answer=row["source_answer"],
                original_review_id=row["review_id"],
                original_class=row["proposed_class"],
                original_status=row["review_status"],
                original_label=row["review_label"],
                original_notes=row["review_notes"],
                original_outcome=f"V2_{row['review_status']}",
                eligibility_note=(
                    "The prior STOP rejection is not reused as approval. The image "
                    "visually depicts a bus-stop or bus-bay sign and requires a new decision."
                ),
                inventory=inventory,
                groups=groups,
                split_evidence=split_evidence,
                all_evidence=all_evidence,
            )
        )

    for source_id, (new_label, eligibility_note) in LOCAL_ELIGIBILITY.items():
        source = excluded[source_id]
        candidates.append(
            _make_row(
                source_id=source_id,
                new_label=new_label,
                source_question=source["source_question"],
                source_answer=source["source_answer"],
                original_review_id="",
                original_class=source["proposed_label"],
                original_status="excluded_before_v2_review",
                original_label="",
                original_notes="",
                original_outcome=source["reason"],
                eligibility_note=eligibility_note,
                inventory=inventory,
                groups=groups,
                split_evidence=split_evidence,
                all_evidence=all_evidence,
            )
        )

    class_order = {label: index for index, label in enumerate(BUNDLE_CLASS_ORDER)}
    candidates.sort(
        key=lambda row: (
            class_order[row["proposed_unseen_class"]],
            row["source_image_id"],
        )
    )
    for index, row in enumerate(candidates, 1):
        row["review_id"] = f"ULR-{index:04d}"
    return candidates


def _make_row(
    *,
    source_id: str,
    new_label: str,
    source_question: str,
    source_answer: str,
    original_review_id: str,
    original_class: str,
    original_status: str,
    original_label: str,
    original_notes: str,
    original_outcome: str,
    eligibility_note: str,
    inventory: Mapping[str, Mapping[str, str]],
    groups: Mapping[str, str],
    split_evidence: Mapping[str, Any],
    all_evidence: Any,
) -> dict[str, str]:
    item = inventory[source_id]
    group_id = groups[source_id]
    overlaps: dict[str, str] = {}
    overlap_reasons: list[str] = []
    for split in ("train", "validation", "test"):
        overlap, reason = frozen_split_overlap(
            source_id=source_id,
            perceptual_group_id=group_id,
            sha256=item["sha256"],
            evidence=split_evidence[split],
        )
        overlaps[split] = "yes" if overlap else "no"
        if overlap:
            overlap_reasons.append(f"{split}:{reason}")

    exact_overlap = item["sha256"] in all_evidence.sha256_digests
    group_overlap = group_id in all_evidence.perceptual_group_ids
    source_overlap = source_id in all_evidence.source_ids
    if source_overlap or exact_overlap or group_overlap or overlap_reasons:
        raise LocalUnseenReviewError(
            f"Candidate overlaps frozen V2: {source_id} ({';'.join(overlap_reasons)})"
        )

    dependency_group = group_id
    if source_id in {"img_0786.jpg", "img_0812.jpg"}:
        dependency_group = "manual_source_group_no_parking_roadside_01"
    return {
        "review_id": "",
        "image_path": f"data/raw/indian_traffic_vqa/traffic512final/{source_id}",
        "source_image_id": source_id,
        "source_question": source_question,
        "source_answer": source_answer,
        "source_dataset": "Indian Traffic VQA Dataset B",
        "original_v2_review_id": original_review_id,
        "original_v2_proposed_class": original_class,
        "original_v2_review_status": original_status,
        "original_v2_review_label": original_label,
        "original_v2_review_notes": original_notes,
        "original_v2_review_outcome": original_outcome,
        "proposed_unseen_class": new_label,
        "sha256": item["sha256"],
        "dhash": item["difference_hash"],
        "perceptual_group_id": group_id,
        "source_or_independence_group": dependency_group,
        "v2_train_overlap": overlaps["train"],
        "v2_validation_overlap": overlaps["validation"],
        "v2_test_overlap": overlaps["test"],
        "v2_exact_sha_overlap": "yes" if exact_overlap else "no",
        "v2_perceptual_group_overlap": "yes" if group_overlap else "no",
        "overlap_validation_notes": (
            "zero source-image, perceptual-group, and exact-SHA overlap with frozen V2"
        ),
        "eligibility_note": eligibility_note,
        "licence_status": "pending_curator_confirmation",
        "review_status": "pending",
        "review_label": "",
        "review_notes": "",
    }


def _write_csv(path: Path, rows: Sequence[Mapping[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS, extrasaction="raise")
        writer.writeheader()
        writer.writerows(rows)


def _write_workbook(path: Path, rows: Sequence[Mapping[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pending Review"
    sheet.append(list(COLUMNS))
    for row in rows:
        sheet.append([row[column] for column in COLUMNS])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(COLUMNS, 1):
        values = [column, *(str(row[column]) for row in rows)]
        width = min(max(len(value) for value in values) + 2, 55)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 12)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)


def _write_pdf(path: Path, rows: Sequence[Mapping[str, str]]) -> None:
    styles = getSampleStyleSheet()
    heading = ParagraphStyle(
        "PendingHeading",
        parent=styles["Heading1"],
        textColor=colors.HexColor("#9C0006"),
        fontSize=15,
        leading=18,
        alignment=1,
        spaceAfter=3 * mm,
    )
    subheading = ParagraphStyle(
        "ClassHeading",
        parent=styles["Heading2"],
        fontSize=11,
        leading=13,
        alignment=1,
        spaceAfter=3 * mm,
    )
    caption = ParagraphStyle(
        "ReviewCaption",
        parent=styles["BodyText"],
        fontSize=7.5,
        leading=9.2,
        spaceAfter=0,
    )
    document = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A4),
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
        title="Local unseen-class pending human review bundle",
        author="Adaptive Indian Road Sign Recognition project",
    )
    pages: list[tuple[str, Sequence[Mapping[str, str]]]] = []
    for label in BUNDLE_CLASS_ORDER:
        class_rows = [row for row in rows if row["proposed_unseen_class"] == label]
        for start in range(0, len(class_rows), 2):
            pages.append((label, class_rows[start : start + 2]))

    story: list[Any] = []
    class_page_counts: dict[str, int] = {}
    class_totals = {
        label: (len([row for row in rows if row["proposed_unseen_class"] == label]) + 1)
        // 2
        for label in BUNDLE_CLASS_ORDER
    }
    for page_index, (label, page_rows) in enumerate(pages):
        class_page_counts[label] = class_page_counts.get(label, 0) + 1
        story.append(
            Paragraph("PENDING HUMAN REVIEW — NOT APPROVED FOR EXPERIMENT", heading)
        )
        story.append(
            Paragraph(
                f"Proposed unseen class: {_escape(label)} | "
                f"class page {class_page_counts[label]} of {class_totals[label]}",
                subheading,
            )
        )
        cards: list[Any] = []
        for row in page_rows:
            image_path = PROJECT_ROOT / row["image_path"]
            cards.append(
                [
                    _scaled_image(image_path, 126 * mm, 88 * mm),
                    Spacer(1, 2 * mm),
                    Paragraph(_caption_html(row), caption),
                ]
            )
        while len(cards) < 2:
            cards.append(Paragraph("", caption))
        table = Table([cards], colWidths=[138 * mm, 138 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#808080")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFBFBF")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4 * mm),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4 * mm),
                    ("TOPPADDING", (0, 0), (-1, -1), 3 * mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3 * mm),
                ]
            )
        )
        story.append(table)
        if page_index != len(pages) - 1:
            story.append(PageBreak())
    document.build(story)


def _caption_html(row: Mapping[str, str]) -> str:
    values = (
        ("review_id", row["review_id"]),
        ("proposed unseen class", row["proposed_unseen_class"]),
        ("source image ID", row["source_image_id"]),
        ("original V2 proposed class", row["original_v2_proposed_class"]),
        ("original V2 review status", row["original_v2_review_status"]),
        ("conservative/perceptual group", row["source_or_independence_group"]),
        ("eligibility note", row["eligibility_note"]),
    )
    return "<br/>".join(
        f"<b>{_escape(label)}:</b> {_escape(value)}" for label, value in values
    )


def _scaled_image(path: Path, maximum_width: float, maximum_height: float) -> Image:
    if not path.is_file():
        raise LocalUnseenReviewError(f"Review image is missing: {path}")
    with PilImage.open(path) as source:
        width, height = source.size
        source.verify()
    scale = min(maximum_width / width, maximum_height / height)
    return Image(str(path), width=width * scale, height=height * scale)


def _validate_artifacts(
    pdf_path: Path,
    xlsx_path: Path,
    csv_path: Path,
    rows: Sequence[Mapping[str, str]],
) -> int:
    expected_ids = [row["review_id"] for row in rows]
    csv_rows = _read_csv(csv_path)
    if [row["review_id"] for row in csv_rows] != expected_ids:
        raise LocalUnseenReviewError("CSV review IDs differ from source rows")
    validate_local_unseen_review_rows(csv_rows)

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    if workbook.sheetnames != ["Pending Review"]:
        raise LocalUnseenReviewError(f"Unexpected XLSX sheets: {workbook.sheetnames}")
    sheet = workbook["Pending Review"]
    headers = [str(cell.value) for cell in sheet[1]]
    if headers != list(COLUMNS):
        raise LocalUnseenReviewError("XLSX columns differ from locked schema")
    xlsx_ids = [
        str(sheet.cell(index, 1).value) for index in range(2, sheet.max_row + 1)
    ]
    workbook.close()
    if xlsx_ids != expected_ids or len(xlsx_ids) != 43:
        raise LocalUnseenReviewError("XLSX review IDs differ from source rows")

    reader = PdfReader(str(pdf_path))
    pdf_text = "\n".join(page.extract_text() or "" for page in reader.pages)
    pdf_ids = re.findall(r"ULR-\d{4}", pdf_text)
    if pdf_ids != expected_ids:
        raise LocalUnseenReviewError(
            "PDF review IDs do not exactly match the XLSX/CSV order"
        )
    if pdf_text.count("PENDING HUMAN REVIEW") != len(reader.pages):
        raise LocalUnseenReviewError("Required pending-review heading is missing")
    if len(reader.pages) != 22:
        raise LocalUnseenReviewError(
            f"Expected 22 readable PDF pages, found {len(reader.pages)}"
        )
    return len(reader.pages)


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise LocalUnseenReviewError(f"CSV has no header: {path}")
        return [dict(row) for row in reader]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _escape(value: str) -> str:
    return value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


if __name__ == "__main__":
    raise SystemExit(main())
